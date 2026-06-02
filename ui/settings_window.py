"""
ui/settings_window.py

设置界面 —— 独立窗口，标签页布局 / Settings Interface — Independent Window, Tabbed Layout:
  1. 通用设置（全局监听开关，界面语言等） / General Settings (Global listener, interface language, etc.)
  2. 映射组管理（创建、编辑、删除映射组，快捷键绑定与启用状态，支持拖拽调节宽度） / Mapping Groups (Create, edit, delete mapping groups, hotkey bindings, draggable list width)
  3. 查询模块管理（添加、移除、配置、暂停/启用后端模块） / Query Modules (Add, remove, configure, pause/enable backend modules)
  4. 交互模块管理（创建、编辑、删除交互模块） / Interaction Modules (Create, edit, delete custom dialog layouts)
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING, Any

from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtWidgets import (
    QDialog, QTabWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QWidget, QLabel, QLineEdit, QCheckBox, QSpinBox, QDoubleSpinBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QPushButton, QGroupBox, QComboBox, QMessageBox, QSplitter,
    QListWidget, QListWidgetItem, QFrame, QScrollArea, QSizePolicy,
)

from core.mapping import (
    InteractionModule, BASE_TYPES, BASE_TYPE_LABELS,
    ComponentDef, ConfigManager,
)
from core.i18n import _t
from ui.display_panel import ALL_COMP_TYPES, COMP_TYPE_LABELS

if TYPE_CHECKING:
    from main import Application

logger = logging.getLogger(__name__)


def _ui_index_to_col_letter(idx: int) -> str:
    """0-based 整数 → Excel 列字母（UI 层工具函数）。0→A, 25→Z, 26→AA"""
    result = ""
    n = idx + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result



# ══════════════════════════════════════════════════════════════════
# 后台连接测试线程
# ══════════════════════════════════════════════════════════════════

class _ConnectTestThread(QThread):
    """在后台执行 backend.connect()，避免阻塞 UI"""
    result = pyqtSignal(bool, str)   # (success, message)

    def __init__(self, backend_cls, config: dict, parent=None):
        super().__init__(parent)
        self._cls = backend_cls
        self._cfg = config

    def run(self):
        try:
            instance = self._cls.from_config(self._cfg)
            ok = instance.connect()
            if ok:
                self.result.emit(True, "连接成功，服务就绪")
            else:
                status = instance.get_status()
                self.result.emit(False, f"服务未就绪（状态: {status}）")
        except Exception as exc:
            self.result.emit(False, str(exc)[:120])


# ══════════════════════════════════════════════════════════════════
# 设置主窗口
# ══════════════════════════════════════════════════════════════════

class SettingsWindow(QDialog):
    """设置界面主窗口"""

    # 配置变更信号
    config_changed = pyqtSignal()

    def __init__(self, app_controller: Application, parent=None):
        super().__init__(parent, Qt.Window)
        self._app = app_controller
        self.setWindowTitle(_t("Excel 快捷查询助手 — 设置"))
        self.setMinimumSize(820, 580)

        layout = QVBoxLayout(self)
        self._tabs = QTabWidget()

        self._general_tab = _GeneralTab(app_controller)
        self._mapping_tab = _MappingTab(app_controller)
        self._qm_tab = _QueryModuleTab(app_controller)
        self._im_tab = _InteractionModuleTab(app_controller)

        self._tabs.addTab(self._general_tab, _t("通用设置"))
        self._tabs.addTab(self._mapping_tab, _t("映射组管理"))
        self._tabs.addTab(self._qm_tab, _t("查询模块管理"))
        self._tabs.addTab(self._im_tab, _t("交互模块管理"))

        layout.addWidget(self._tabs)

        # 底部按钮
        btn_bar = QHBoxLayout()
        btn_bar.addStretch()
        save_btn = QPushButton(_t("保存并关闭"))
        save_btn.setFixedWidth(120)
        save_btn.clicked.connect(self._save_and_close)
        btn_bar.addWidget(save_btn)
        layout.addLayout(btn_bar)

    def _save_and_close(self) -> None:
        # 强制提交所有表格单元格的当前编辑
        self._mapping_tab._table.setCurrentCell(-1, -1)
        self._mapping_tab._table.clearFocus()
        self._general_tab.apply()
        try:
            self._mapping_tab.apply()
        except ValueError:
            return  # 校验失败，拦截保存并阻断窗口关闭
        self._app.config_mgr.save()
        self.config_changed.emit()
        self.close()

    def showEvent(self, event):
        """每次显示时刷新数据"""
        self._general_tab.refresh()
        self._mapping_tab.refresh()
        self._qm_tab.refresh()
        self._im_tab.refresh()
        super().showEvent(event)


# ══════════════════════════════════════════════════════════════════
# Tab 1: 通用设置（已简化，查询模块参数移至专属 Tab）
# ══════════════════════════════════════════════════════════════════

class _GeneralTab(QWidget):
    def __init__(self, app_controller: Application):
        super().__init__()
        self._app = app_controller
        layout = QVBoxLayout(self)

        # 全局开关
        self._hotkey_cb = QCheckBox(_t("启用全局键盘监听（勾选后主程序才能通过快捷键唤起查询面板）"))
        self._hotkey_cb.setStyleSheet("font-size: 13px;")
        layout.addWidget(self._hotkey_cb)

        # 界面语言设置
        lang_gb = QGroupBox(_t("界面语言 (Language):"))
        lang_f = QFormLayout(lang_gb)
        self._lang_combo = QComboBox()
        self._lang_combo.addItem(_t("跟随系统 (Auto)"), "auto")
        self._lang_combo.addItem(_t("中文 (简体)"), "zh")
        self._lang_combo.addItem(_t("English"), "en")
        lang_f.addRow(_t("选择语言 (Select Language):"), self._lang_combo)
        layout.addWidget(lang_gb)

        # 提示跳转
        note = QLabel("💡 " + _t("查询模块的连接参数（地址、数据库、超时等）\n   请在「查询模块」标签页中配置。"))
        note.setStyleSheet("color: #666; font-size: 12px; padding: 8px;")
        layout.addWidget(note)

        layout.addStretch()

    def refresh(self) -> None:
        self._hotkey_cb.setChecked(self._app.config_mgr.is_hotkey_enabled())
        # 从配置中获取语言
        lang = self._app.config_mgr.config.get("global", {}).get("language", "auto")
        idx = self._lang_combo.findData(lang)
        if idx >= 0:
            self._lang_combo.setCurrentIndex(idx)

    def apply(self) -> None:
        self._app.config_mgr.config.setdefault("global", {})["hotkey_enabled"] = (
            self._hotkey_cb.isChecked()
        )
        # 保存语言设置
        lang = self._lang_combo.currentData() or "auto"
        self._app.config_mgr.config.setdefault("global", {})["language"] = lang
        self._app.config_mgr.save()
        
        # 动态更新 i18n 中的语言
        from core.i18n import set_language, detect_system_lang
        if lang == "auto":
            set_language(detect_system_lang())
        else:
            set_language(lang)
            
        QMessageBox.information(
            self, 
            _t("提示") if "en" != lang else "Tip", 
            _t("配置已保存。界面语言设置将在重启应用后完全生效。") if "en" != lang else "Settings saved. Interface language changes will fully take effect after restarting the application."
        )







# ══════════════════════════════════════════════════════════════════
# Tab 3: 映射组管理
# ══════════════════════════════════════════════════════════════════

class _MappingTab(QWidget):
    def __init__(self, app_controller: Application):
        super().__init__()
        self._app = app_controller

        layout = QHBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)

        self._splitter = QSplitter(Qt.Horizontal)

        # 左侧：列表与按钮表单区
        left_widget = QWidget()
        left = QVBoxLayout(left_widget)
        left.setContentsMargins(0, 0, 0, 0)
        left.setSpacing(6)

        left.addWidget(QLabel(_t("映射组列表 (右键可基于此创建新组):")))
        
        # 升级为三列表格：映射组名称 | 触发快捷键 | 是否启用
        self._table = QTableWidget()
        self._table.setColumnCount(3)
        self._table.setHorizontalHeaderLabels([_t("映射组名称"), _t("触发快捷键"), _t("是否启用")])
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.Fixed)
        hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        self._table.setColumnWidth(1, 120)
        self._table.setColumnWidth(2, 60)
        self._table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.verticalHeader().setDefaultSectionSize(26)
        left.addWidget(self._table)

        # 信号绑定
        self._table.currentCellChanged.connect(self._on_current_cell_changed)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_context_menu)
        self._table.itemChanged.connect(self._on_item_changed)

        btn_bar = QHBoxLayout()
        add_btn = QPushButton(_t("➕ 新建"))
        add_btn.clicked.connect(self._add_group)
        edit_btn = QPushButton(_t("✏ 编辑"))
        edit_btn.clicked.connect(self._edit_group)
        del_btn = QPushButton(_t("🗑 删除"))
        del_btn.clicked.connect(self._delete_group)

        up_btn = QPushButton("↑")
        up_btn.setFixedWidth(30)
        up_btn.clicked.connect(lambda: self._move_group(-1))

        dn_btn = QPushButton("↓")
        dn_btn.setFixedWidth(30)
        dn_btn.clicked.connect(lambda: self._move_group(1))

        btn_bar.addWidget(add_btn)
        btn_bar.addWidget(edit_btn)
        btn_bar.addWidget(del_btn)
        btn_bar.addSpacing(8)
        btn_bar.addWidget(up_btn)
        btn_bar.addWidget(dn_btn)
        left.addLayout(btn_bar)

        # 右侧：详情预览（加入滚动区域以防溢出）
        right_widget = QWidget()
        right = QVBoxLayout(right_widget)
        right.setContentsMargins(0, 0, 0, 0)
        right.setSpacing(6)
        right.addWidget(QLabel(_t("映射组详情:")))

        from PyQt5.QtWidgets import QScrollArea, QFrame
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.StyledPanel)
        scroll.setStyleSheet("background: #f5f5f5; border-radius: 4px; border: 1px solid #e0e0e0;")

        self._preview = QLabel(_t("（请选择映射组）"))
        self._preview.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self._preview.setWordWrap(True)
        self._preview.setStyleSheet("padding: 10px; background: transparent;")
        scroll.setWidget(self._preview)
        right.addWidget(scroll, 1)

        # 将左右面板并入 Splitter
        self._splitter.addWidget(left_widget)
        self._splitter.addWidget(right_widget)
        self._splitter.setStretchFactor(0, 3)
        self._splitter.setStretchFactor(1, 4)

        # 还原上次拖拽的尺寸
        saved_sizes = self._app.config_mgr.get_splitter_sizes("mapping_splitter")
        if saved_sizes:
            self._splitter.setSizes(saved_sizes)
        else:
            self._splitter.setSizes([350, 450])

        layout.addWidget(self._splitter)

    # ── 数据刷新与载入 ─────────────────────────────────────────────

    def refresh(self) -> None:
        self._table.blockSignals(True)
        self._table.setRowCount(0)

        cfg = self._app.config_mgr
        groups = cfg.get_mapping_groups()
        tm = cfg.get_table_modules()

        # 构建快捷键映射表：mg_id -> [(mod_id, mod_cfg)]
        mg_hotkeys = {}
        for mod_id, mod_cfg in tm.items():
            mg_id = mod_cfg.get("mapping_group", "")
            if mg_id:
                mg_hotkeys.setdefault(mg_id, []).append((mod_id, mod_cfg))

        for gid, g in groups.items():
            row = self._table.rowCount()
            self._table.insertRow(row)

            # 列 0: 映射组名称 (只读)
            name_item = QTableWidgetItem(g.name)
            name_item.setData(Qt.UserRole, gid)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)

            # 查询关联快捷键信息
            hotkey_str = ""
            enabled = False
            associated_mod_id = ""

            hk_list = mg_hotkeys.get(gid, [])
            if hk_list:
                # 优先显示启用的快捷键，否则显示第一个
                hk_list_sorted = sorted(hk_list, key=lambda x: 1 if x[1].get("enabled", True) else 0, reverse=True)
                mod_id, mod_cfg = hk_list_sorted[0]
                hotkey_str = mod_cfg.get("hotkey", "")
                enabled = mod_cfg.get("enabled", True)
                associated_mod_id = mod_id

            name_item.setData(Qt.UserRole + 1, associated_mod_id)
            self._table.setItem(row, 0, name_item)

            # 列 1: 触发快捷键 (双击可修改)
            hk_item = QTableWidgetItem(hotkey_str)
            self._table.setItem(row, 1, hk_item)

            # 列 2: 是否启用 (复选框，居中对齐)
            cb = QCheckBox()
            cb.setChecked(enabled)
            cb.toggled.connect(lambda checked, cbox=cb: self._on_checkbox_toggled(cbox))

            w = QWidget()
            h = QHBoxLayout(w)
            h.addWidget(cb)
            h.setAlignment(Qt.AlignCenter)
            h.setContentsMargins(0, 0, 0, 0)
            self._table.setCellWidget(row, 2, w)

        self._table.blockSignals(False)
        self._preview.setText(_t("（请选择映射组）"))
        self._table.setCurrentCell(-1, -1)

    # ── 选择联动与预览 ─────────────────────────────────────────────

    def _on_current_cell_changed(self, row: int, col: int, prev_row: int, prev_col: int) -> None:
        if row != prev_row:
            self._on_select(row)

    def _on_select(self, row: int) -> None:
        if row < 0 or row >= self._table.rowCount():
            self._preview.setText(_t("（请选择映射组）"))
            return
        item = self._table.item(row, 0)
        if not item:
            self._preview.setText(_t("（请选择映射组）"))
            return
        gid = item.data(Qt.UserRole)
        group = self._app.config_mgr.get_mapping_group(gid)
        if group is None:
            return

        im_id = group.interaction_module
        im_cfg = self._app.config_mgr.get_interaction_modules().get(im_id) if im_id else None
        im_desc = f"{im_cfg.name} [{im_id}]" if im_cfg else (im_id or _t("（未设置）"))

        qm_id = group.query_module
        qm_cfg = self._app.config_mgr.get_query_modules().get(qm_id) if qm_id else None
        qm_desc = f"{qm_cfg.get('name', qm_id)} [{qm_id}]" if qm_cfg else (qm_id or _t("（未设置）"))

        lines = [
            _t("<b>名称:</b> ") + group.name,
            _t("<b>交互模块:</b> ") + im_desc,
            _t("<b>查询模块:</b> ") + qm_desc,
            _t("<b>查询参数:</b> ") + f"limit={group.query_params.get('limit', 20)}",
            "",
            _t("<b>输入映射（Excel → 查询字段）:</b>"),
        ]

        for im in group.input_mappings:
            target_desc = im.query_field if im.query_field else _t("【提示: {comp}】").format(comp=im.ui_component)
            def_desc = _t("（默认: '{val}'）").format(val=im.default_value) if getattr(im, "default_value", "") else ""
            lines.append(_t("  • 列 {cols} → {target_desc}（分隔: '{sep}'）{def_desc}").format(
                cols=', '.join(im.columns), target_desc=target_desc, sep=im.separator, def_desc=def_desc
            ))

        lines.append("")
        lines.append(_t("<b>显示列（弹窗表格）:</b>"))
        for dc in group.display_columns:
            lines.append(_t("  • {field} → 「{header}」宽 {width}px").format(
                field=dc.field, header=dc.header, width=dc.width
            ))
        lines.append("")
        lines.append(_t("<b>输出映射（结果 → Excel）:</b>"))
        for om in group.output_mappings:
            lines.append(_t("  • {field} → 列 {col}").format(
                field=om.result_field, col=om.column
            ))
        self._preview.setText("<br>".join(lines))

    # ── 新建与编辑 ──────────────────────────────────────────────────

    def _add_group(self) -> None:
        dlg = MappingGroupDialog(self._app, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            gid, group = dlg.get_result()
            self._app.config_mgr.save_mapping_group(gid, group)
            self.refresh()

    def _edit_group(self) -> None:
        row = self._table.currentRow()
        if row < 0:
            QMessageBox.information(self, _t("提示"), _t("请先选择要编辑的映射组"))
            return
        gid = self._table.item(row, 0).data(Qt.UserRole)
        group = self._app.config_mgr.get_mapping_group(gid)
        if group is None:
            return
        dlg = MappingGroupDialog(self._app, group_id=gid, group=group, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            new_gid, new_group = dlg.get_result()
            if new_gid != gid:
                self._app.config_mgr.delete_mapping_group(gid)
            self._app.config_mgr.save_mapping_group(new_gid, new_group)
            self.refresh()

    def _delete_group(self) -> None:
        row = self._table.currentRow()
        if row < 0:
            return
        gid = self._table.item(row, 0).data(Qt.UserRole)
        mg = self._app.config_mgr.get_mapping_group(gid)
        display = f"{mg.name} [{gid}]" if mg else gid
        ret = QMessageBox.question(self, _t("确认删除"), _t("确定删除映射组「{display}」？").format(display=display))
        if ret == QMessageBox.Yes:
            self._app.config_mgr.delete_mapping_group(gid)
            self.refresh()

    def _move_group(self, direction: int) -> None:
        row = self._table.currentRow()
        if row < 0:
            return
        gid = self._table.item(row, 0).data(Qt.UserRole)
        success = self._app.config_mgr.move_mapping_group(gid, direction)
        if success:
            self.refresh()
            new_row = row + direction
            if 0 <= new_row < self._table.rowCount():
                self._table.setCurrentCell(new_row, 0)

    # ── 右键复制创建 ─────────────────────────────────────────────────

    def _on_context_menu(self, pos) -> None:
        item = self._table.itemAt(pos)
        if item is None:
            return
        row = self._table.row(item)
        name_item = self._table.item(row, 0)
        if name_item is None:
            return

        from PyQt5.QtWidgets import QMenu
        menu = QMenu(self)
        dup_action = menu.addAction(_t("➕ 基于此创建新映射组"))

        action = menu.exec_(self._table.mapToGlobal(pos))
        if action == dup_action:
            self._duplicate_group(name_item.data(Qt.UserRole))

    def _duplicate_group(self, gid: str) -> None:
        group = self._app.config_mgr.get_mapping_group(gid)
        if group is None:
            return

        import copy
        new_group = copy.deepcopy(group)
        new_group.name = _t("复制自-{name}").format(name=new_group.name)

        dlg = MappingGroupDialog(self._app, group_id="", group=new_group, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            new_gid, saved_group = dlg.get_result()
            self._app.config_mgr.save_mapping_group(new_gid, saved_group)
            self.refresh()

    # ── 快捷键防冲突动态校验 ──────────────────────────────────────────

    def _find_widget_row(self, widget: QWidget) -> int:
        for r in range(self._table.rowCount()):
            w2 = self._table.cellWidget(r, 2)
            if w2 and (w2 == widget or w2.findChild(QCheckBox) == widget):
                return r
        return -1

    def _on_checkbox_toggled(self, cb: QCheckBox) -> None:
        row = self._find_widget_row(cb)
        if row >= 0:
            self._validate_hotkeys(row)

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() == 1:  # 快捷键文本框修改
            self._validate_hotkeys(item.row())

    def _validate_hotkeys(self, trigger_row: int) -> None:
        """
        核心防冲突验证：相同快捷键可以存在，但无法同时处于启用状态。
        """
        hk_item = self._table.item(trigger_row, 1)
        hotkey = hk_item.text().strip() if hk_item else ""
        if not hotkey:
            return

        # 检查触发行的复选框启用状态
        w = self._table.cellWidget(trigger_row, 2)
        cb = w.findChild(QCheckBox) if w else None
        if not cb or not cb.isChecked():
            return  # 没启用，不需要做冲突检测

        # 检查是否与其它已启用的快捷键产生冲突
        for r in range(self._table.rowCount()):
            if r == trigger_row:
                continue

            other_w = self._table.cellWidget(r, 2)
            other_cb = other_w.findChild(QCheckBox) if other_w else None
            if other_cb and other_cb.isChecked():
                other_hk_item = self._table.item(r, 1)
                other_hk = other_hk_item.text().strip() if other_hk_item else ""
                if other_hk.lower() == hotkey.lower():
                    # 发现严重冲突！
                    name_item = self._table.item(r, 0)
                    lbl = name_item.text().strip() if name_item else f"{_t('行')} {r+1}"

                    QMessageBox.warning(
                        self, _t("快捷键冲突"),
                        _t("快捷键「{hotkey}」已被启用的模块「{lbl}」占用。\n"
                           "同一快捷键组合不能同时开启启用状态！").format(hotkey=hotkey, lbl=lbl)
                    )

                    # 撤销本次勾选（临时遮蔽信号防止回环触发）
                    cb.blockSignals(True)
                    cb.setChecked(False)
                    cb.blockSignals(False)
                    break

    # ── 合并应用保存 ────────────────────────────────────────────────

    def apply(self) -> None:
        self._table.model().submit()

        cfg = self._app.config_mgr
        original_tm = cfg.get_table_modules()
        new_tm = {}
        used_keys = set()

        # 1. 最终保存前做一遍全局的双重启用校验
        enabled_hotkeys = {}
        for row in range(self._table.rowCount()):
            w = self._table.cellWidget(row, 2)
            cb = w.findChild(QCheckBox) if w else None
            enabled = cb.isChecked() if cb else True

            hk_item = self._table.item(row, 1)
            hotkey = hk_item.text().strip() if hk_item else ""

            name_item = self._table.item(row, 0)
            name = name_item.text().strip() if name_item else f"行 {row+1}"

            if enabled and hotkey:
                lower_hk = hotkey.lower()
                if lower_hk in enabled_hotkeys:
                    QMessageBox.critical(
                        self, _t("保存失败"),
                        _t("映射组「{lbl}」与「{other_lbl}」配置了相同的已启用快捷键「{hotkey}」！\n"
                           "请先关闭其中一方的启用复选框，再尝试保存。").format(
                               lbl=name, other_lbl=enabled_hotkeys[lower_hk], hotkey=hotkey
                           )
                    )
                    raise ValueError("快捷键保存冲突校验未通过")
                enabled_hotkeys[lower_hk] = name

        # 2. 保留不属于当前映射组（可能手动配置）的外置模块
        for mod_id, mod_cfg in original_tm.items():
            mg_id = mod_cfg.get("mapping_group", "")
            if not mg_id or mg_id not in cfg.get_mapping_groups():
                new_tm[mod_id] = mod_cfg
                used_keys.add(mod_id)

        # 3. 将本表格编辑的数据同步回写到 table_modules 段中
        for row in range(self._table.rowCount()):
            name_item = self._table.item(row, 0)
            if name_item is None:
                continue
            mg_id = name_item.data(Qt.UserRole)
            orig_mod_id = name_item.data(Qt.UserRole + 1)
            name = name_item.text().strip()

            hk_item = self._table.item(row, 1)
            hotkey = hk_item.text().strip() if hk_item else ""

            w = self._table.cellWidget(row, 2)
            cb = w.findChild(QCheckBox) if w else None
            enabled = cb.isChecked() if cb else False

            # 空快捷键，表示不设置，垃圾自动回收清空
            if not hotkey:
                continue

            # 确定 mod_id 唯一主键
            if orig_mod_id:
                mod_id = orig_mod_id
            else:
                mod_id = hotkey.replace("+", "_").replace(" ", "_")
                if not mod_id:
                    mod_id = cfg.generate_id("hotkey_")

            base_key = mod_id
            suffix = 2
            while mod_id in used_keys:
                mod_id = f"{base_key}_{suffix}"
                suffix += 1
            used_keys.add(mod_id)

            entry = {
                "hotkey": hotkey,
                "mapping_group": mg_id,
                "enabled": enabled,
                "title": name,
            }
            new_tm[mod_id] = entry

        cfg.config["table_modules"] = new_tm

        # 同步保存 splitter 分割条尺寸
        cfg.save_splitter_sizes("mapping_splitter", self._splitter.sizes())
        logger.info("已将映射组所配置的快捷键保存到配置，共 %d 条", len(new_tm))


# ══════════════════════════════════════════════════════════════════
# Tab 4: 查询模块管理（新增）
# ══════════════════════════════════════════════════════════════════

class _QueryModuleTab(QWidget):
    """
    查询模块管理标签页。 / Query Module Management Tab.
    支持：添加、移除、配置（含专项与通用键值对兜底）、暂停/启用、测试连接。 / Supports: Add, remove, configure (with backup parameters), pause/enable, and connection testing.
    """

    # 已有专项配置表单的后端类型
    _SPECIALIZED_TYPES: frozenset[str] = frozenset({
        "excel_file", "qdrant_vector", "sql_table",
    })
    # 与 vector_api 共用 base_url+timeout 表单的后端类型
    _URL_TYPES: frozenset[str] = frozenset()

    def __init__(self, app_controller: Application):
        super().__init__()
        self._app = app_controller
        self._current_mod_id: str | None = None
        self._is_new: bool = False
        self._test_thread: _ConnectTestThread | None = None
        self._setup_ui()

    # ── UI 构建 ──────────────────────────────────────────────────

    def _setup_ui(self) -> None:
        outer = QHBoxLayout(self)
        outer.setContentsMargins(6, 6, 6, 6)
        outer.setSpacing(0)

        # ── 左侧：列表 + 操作按钮 ─────────────────────────────────
        left_widget = QWidget()
        left_widget.setFixedWidth(222)
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 6, 0)
        left_layout.setSpacing(4)

        left_layout.addWidget(QLabel(_t("查询模块列表:")))

        self._list = QListWidget()
        self._list.setSpacing(2)
        self._list.setStyleSheet("""
            QListWidget::item { padding: 4px 6px; border-radius: 4px; }
            QListWidget::item:selected { background: #e3f2fd; color: #0d47a1; }
        """)
        self._list.currentRowChanged.connect(self._on_select)
        left_layout.addWidget(self._list)

        btn_bar = QHBoxLayout()
        btn_bar.setSpacing(4)

        self._add_btn = QPushButton(_t("➕ 新建"))
        self._add_btn.setToolTip(_t("新建查询模块"))
        self._add_btn.clicked.connect(self._on_add)

        self._del_btn = QPushButton(_t("🗑 删除"))
        self._del_btn.setToolTip(_t("删除选中模块"))
        self._del_btn.clicked.connect(self._on_delete)

        self._toggle_btn = QPushButton(_t("⏸ 暂停"))
        self._toggle_btn.setToolTip(_t("暂停/启用选中模块"))
        self._toggle_btn.setEnabled(False)
        self._toggle_btn.clicked.connect(self._on_toggle)

        btn_bar.addWidget(self._add_btn)
        btn_bar.addWidget(self._del_btn)
        btn_bar.addWidget(self._toggle_btn)
        btn_bar.addStretch()
        left_layout.addLayout(btn_bar)

        # ── 竖向分隔线 ────────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setStyleSheet("color: #e0e0e0;")

        # ── 右侧：配置面板（滚动区域） ─────────────────────────────
        right_widget = QWidget()
        right_outer = QVBoxLayout(right_widget)
        right_outer.setContentsMargins(8, 0, 0, 0)
        right_outer.setSpacing(0)

        # 占位提示（未选中时显示）
        self._placeholder = QLabel(_t("← 请从左侧选择一个模块，\n   或点击「➕ 新建」添加模块"))
        self._placeholder.setAlignment(Qt.AlignCenter)
        self._placeholder.setStyleSheet("color: #9e9e9e; font-size: 13px;")
        right_outer.addWidget(self._placeholder, 1)

        # 表单容器（选中后显示）
        self._form_scroll = QScrollArea()
        self._form_scroll.setWidgetResizable(True)
        self._form_scroll.setFrameShape(QFrame.NoFrame)
        self._form_inner = QWidget()
        self._build_form(self._form_inner)
        self._form_scroll.setWidget(self._form_inner)
        self._form_scroll.hide()
        right_outer.addWidget(self._form_scroll, 1)

        outer.addWidget(left_widget)
        outer.addWidget(sep)
        outer.addWidget(right_widget, 1)

    def _build_form(self, parent: QWidget) -> None:
        fl = QVBoxLayout(parent)
        fl.setContentsMargins(4, 4, 12, 4)
        fl.setSpacing(10)

        # ── 基本信息 ─────────────────────────────────────────────
        basic_gb = QGroupBox(_t("基本信息"))
        bf = QFormLayout(basic_gb)
        bf.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        bf.setSpacing(6)

        self._id_edit = QLineEdit()
        self._id_edit.setPlaceholderText(_t("模块唯一标识（英文字母 + 下划线）"))
        bf.addRow(_t("模块 ID:"), self._id_edit)

        self._name_edit = QLineEdit()
        self._name_edit.setPlaceholderText(_t("可读名称，用于列表显示（例如：PG物料精确查询）"))
        bf.addRow(_t("模块名称:"), self._name_edit)

        self._type_combo = QComboBox()
        self._type_combo.setMinimumWidth(200)
        self._type_combo.currentIndexChanged.connect(self._on_type_changed)
        bf.addRow(_t("后端类型:"), self._type_combo)

        self._enabled_cb = QCheckBox(_t("启用此模块（取消勾选则跳过加载，快捷键触发时会提示暂停）"))
        self._enabled_cb.setChecked(True)
        self._enabled_cb.setStyleSheet("font-size: 12px;")
        bf.addRow("", self._enabled_cb)

        fl.addWidget(basic_gb)

        # ── vector_api / pg_material_api 专项参数 ───────────────
        self._vector_gb = QGroupBox(_t("API 服务参数（vector_api / pg_material_api）"))
        vf = QFormLayout(self._vector_gb)
        vf.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        vf.setSpacing(6)

        self._v_url = QLineEdit()
        self._v_url.setPlaceholderText("http://host:port")
        vf.addRow(_t("API 地址:"), self._v_url)

        self._v_timeout = QSpinBox()
        self._v_timeout.setRange(5, 300)
        self._v_timeout.setValue(30)
        self._v_timeout.setSuffix(_t(" 秒"))
        vf.addRow(_t("超时:"), self._v_timeout)

        fl.addWidget(self._vector_gb)

        # ── PostgreSQL 专项参数 ──────────────────────────────────
        self._pg_gb = QGroupBox(_t("PostgreSQL 数据库参数（pg_plan / pg_spare_parts）"))
        pf = QFormLayout(self._pg_gb)
        pf.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        pf.setSpacing(6)

        self._pg_host = QLineEdit()
        self._pg_host.setPlaceholderText(_t("数据库主机地址"))
        pf.addRow(_t("主机:"), self._pg_host)

        self._pg_port = QSpinBox()
        self._pg_port.setRange(1, 65535)
        self._pg_port.setValue(5432)
        pf.addRow(_t("端口:"), self._pg_port)

        self._pg_dbname = QLineEdit()
        self._pg_dbname.setPlaceholderText(_t("数据库名称"))
        pf.addRow(_t("数据库:"), self._pg_dbname)

        self._pg_user = QLineEdit()
        self._pg_user.setPlaceholderText(_t("用户名"))
        pf.addRow(_t("用户名:"), self._pg_user)

        # 密码行：输入框 + 显示/隐藏切换按钮
        pass_row = QHBoxLayout()
        self._pg_pass = QLineEdit()
        self._pg_pass.setEchoMode(QLineEdit.Password)
        self._pg_pass.setPlaceholderText(_t("密码（明文存储于 config.yaml）"))
        self._show_pass_btn = QPushButton("👁")
        self._show_pass_btn.setFixedWidth(32)
        self._show_pass_btn.setCheckable(True)
        self._show_pass_btn.setToolTip(_t("切换密码可见性"))
        self._show_pass_btn.toggled.connect(
            lambda checked: self._pg_pass.setEchoMode(
                QLineEdit.Normal if checked else QLineEdit.Password
            )
        )
        pass_row.addWidget(self._pg_pass)
        pass_row.addWidget(self._show_pass_btn)
        pf.addRow(_t("密码:"), pass_row)

        self._pg_timeout = QSpinBox()
        self._pg_timeout.setRange(5, 300)
        self._pg_timeout.setValue(15)
        self._pg_timeout.setSuffix(_t(" 秒"))
        pf.addRow(_t("超时:"), self._pg_timeout)

        fl.addWidget(self._pg_gb)

        # ── 通用扩展参数（未知/自定义后端类型兜底）─────────────────
        self._custom_gb = QGroupBox(_t("扩展参数（键值对，用于自定义 / 未来新增的后端类型）"))
        cf = QVBoxLayout(self._custom_gb)
        cf.setSpacing(4)

        note = QLabel(
            _t("以下键值对将原样传入该后端的 from_config(config) 方法。\n"
            "支持字符串值；如需数值/布尔型请在后端 from_config 中自行转换。")
        )
        note.setStyleSheet("color: #666; font-size: 11px;")
        cf.addWidget(note)

        self._kv_table = QTableWidget(0, 2)
        self._kv_table.setHorizontalHeaderLabels([_t("参数名"), _t("参数值")])
        kv_hdr = self._kv_table.horizontalHeader()
        kv_hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        kv_hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        self._kv_table.setMaximumHeight(160)
        self._kv_table.setAlternatingRowColors(True)
        cf.addWidget(self._kv_table)

        kv_btns = QHBoxLayout()
        add_kv = QPushButton(_t("➕ 添加参数"))
        add_kv.setFixedWidth(100)
        add_kv.clicked.connect(lambda: self._kv_add_row())
        del_kv = QPushButton(_t("🗑 删除行"))
        del_kv.setFixedWidth(80)
        del_kv.clicked.connect(self._kv_del_row)
        kv_btns.addWidget(add_kv)
        kv_btns.addWidget(del_kv)
        kv_btns.addStretch()
        cf.addLayout(kv_btns)

        fl.addWidget(self._custom_gb)

        # ── excel_file 专项参数 ───────────────────────────────
        self._excel_gb = QGroupBox(_t("Excel 文件查询参数（excel_file）"))
        self._excel_gb.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        excel_outer = QVBoxLayout(self._excel_gb)
        excel_outer.setSpacing(8)

        # 数据源
        src_gb = QGroupBox(_t("数据源（文件 / 目录）"))
        src_layout = QVBoxLayout(src_gb)
        src_layout.setSpacing(4)

        self._excel_src_list = QListWidget()
        self._excel_src_list.setMaximumHeight(110)
        self._excel_src_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._excel_src_list.setToolTip(_t("已添加的 Excel 文件 or 目录"))
        self._excel_src_list.itemSelectionChanged.connect(self._on_excel_src_selection_changed)
        src_layout.addWidget(self._excel_src_list)

        src_btns = QHBoxLayout()
        src_btns.setSpacing(4)
        _add_file_btn = QPushButton(_t("📄 添加文件"))
        _add_file_btn.setToolTip(_t("选择一个或多个 Excel 文件"))
        _add_file_btn.clicked.connect(self._excel_add_files)
        _add_dir_btn = QPushButton(_t("📁 添加目录"))
        _add_dir_btn.setToolTip(_t("选择一个目录，自动遍历其中的 Excel 文件"))
        _add_dir_btn.clicked.connect(self._excel_add_directory)
        _del_src_btn = QPushButton(_t("🗑 移除"))
        _del_src_btn.setToolTip(_t("移除选中的数据源"))
        _del_src_btn.clicked.connect(self._excel_del_source)
        src_btns.addWidget(_add_file_btn)
        src_btns.addWidget(_add_dir_btn)
        src_btns.addWidget(_del_src_btn)
        src_btns.addStretch()
        src_layout.addLayout(src_btns)
        excel_outer.addWidget(src_gb)

        # 基本配置
        basic_ef = QFormLayout()
        basic_ef.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        basic_ef.setSpacing(6)

        self._excel_has_header = QCheckBox(_t("第一行为列标题（表头）"))
        self._excel_has_header.setChecked(True)
        basic_ef.addRow("", self._excel_has_header)

        sheet_container = QWidget()
        sheet_layout = QVBoxLayout(sheet_container)
        sheet_layout.setContentsMargins(0, 0, 0, 0)
        sheet_layout.setSpacing(4)
        self._excel_sheet_list = QListWidget()
        self._excel_sheet_list.setMaximumHeight(100)
        self._excel_sheet_list.setSelectionMode(QListWidget.NoSelection)
        self._excel_sheet_list.itemDoubleClicked.connect(self._on_excel_sheet_item_clicked)
        btn_layout = QHBoxLayout()
        self._load_sheets_btn = QPushButton(_t("🔄 加载工作表"))
        self._load_sheets_btn.setToolTip(_t("读取当前选中数据源文件的所有工作表并填入列表"))
        self._load_sheets_btn.clicked.connect(self._excel_load_sheets)
        btn_layout.addWidget(self._load_sheets_btn)
        btn_layout.addStretch()
        sheet_layout.addWidget(self._excel_sheet_list)
        sheet_layout.addLayout(btn_layout)
        basic_ef.addRow(_t("工作表 (可勾选 & 双击预览):"), sheet_container)

        self._excel_max_rows = QSpinBox()
        self._excel_max_rows.setRange(1, 500_000)
        self._excel_max_rows.setValue(100_000)
        self._excel_max_rows.setSuffix(_t(" 行"))
        self._excel_max_rows.setToolTip(_t("单个文件最大加载行数（含表头行）"))
        basic_ef.addRow(_t("最大加载行数:"), self._excel_max_rows)

        excel_outer.addLayout(basic_ef)

        # 列映射
        col_gb = QGroupBox(_t("列映射（留空 = 自动读取全部列）"))
        col_gb.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        col_layout = QVBoxLayout(col_gb)
        col_layout.setSpacing(4)

        col_note = QLabel(
            _t("列标识：可填列字母（A/B/C…）、列序号（1/2/3…）或表头字段名。\n"
            "字段名：该列在查询系统中的名称（供映射组引用）。")
        )
        col_note.setStyleSheet("color:#666; font-size:11px;")
        col_layout.addWidget(col_note)

        self._excel_col_table = QTableWidget(0, 3)
        self._excel_col_table.setHorizontalHeaderLabels([_t("选择"), _t("列标识（字母/序号/字段名）"), _t("字段名（查询用）")])
        self._excel_col_table.setColumnWidth(0, 60)
        ecol_hdr = self._excel_col_table.horizontalHeader()
        ecol_hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        ecol_hdr.setSectionResizeMode(2, QHeaderView.Stretch)
        self._excel_col_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._excel_col_table.setAlternatingRowColors(True)
        col_layout.addWidget(self._excel_col_table)

        col_btns = QHBoxLayout()
        col_btns.setSpacing(4)
        _auto_btn = QPushButton(_t("🔍 自动检测列"))
        _auto_btn.setToolTip(_t("读取第一个数据源文件的首行，自动填充列映射"))
        _auto_btn.clicked.connect(self._excel_autodetect_cols)
        _add_col_btn = QPushButton(_t("➕ 添加行"))
        _add_col_btn.clicked.connect(lambda: self._excel_add_col_row())
        _del_col_btn = QPushButton(_t("🗑 删除行"))
        _del_col_btn.clicked.connect(self._excel_del_col_row)
        col_btns.addWidget(_auto_btn)
        col_btns.addWidget(_add_col_btn)
        col_btns.addWidget(_del_col_btn)
        col_btns.addStretch()
        col_layout.addLayout(col_btns)
        excel_outer.addWidget(col_gb)

        fl.addWidget(self._excel_gb)

        # ── qdrant_vector 专项参数 ─────────────────────────────────
        self._qdrant_gb = QGroupBox(_t("Qdrant 向量数据库参数（qdrant_vector）"))
        self._qdrant_gb.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        qd_outer = QVBoxLayout(self._qdrant_gb)
        qd_outer.setSpacing(8)

        # ── Qdrant 连接 ────────────────────────────────────────────
        qd_conn_gb = QGroupBox(_t("Qdrant 连接"))
        qd_conn_f = QFormLayout(qd_conn_gb)
        qd_conn_f.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        qd_conn_f.setSpacing(5)

        self._qd_host = QLineEdit()
        self._qd_host.setPlaceholderText(_t("IP 地址或域名，如 127.0.0.1"))
        qd_conn_f.addRow(_t("主机 / IP:"), self._qd_host)

        self._qd_port = QSpinBox()
        self._qd_port.setRange(1, 65535)
        self._qd_port.setValue(6334)
        self._qd_port.setToolTip(_t("gRPC 端口通常为 6334，HTTP 端口通常为 6333"))
        qd_conn_f.addRow(_t("端口:"), self._qd_port)

        self._qd_collection = QLineEdit()
        self._qd_collection.setPlaceholderText(_t("collection 名称，如 materials"))
        qd_conn_f.addRow(_t("数据集（Collection）:"), self._qd_collection)

        qd_key_row = QHBoxLayout()
        self._qd_api_key = QLineEdit()
        self._qd_api_key.setEchoMode(QLineEdit.Password)
        self._qd_api_key.setPlaceholderText(_t("API Key（可留空）"))
        _qd_show_btn = QPushButton("👁")
        _qd_show_btn.setFixedWidth(32)
        _qd_show_btn.setCheckable(True)
        _qd_show_btn.toggled.connect(
            lambda c: self._qd_api_key.setEchoMode(
                QLineEdit.Normal if c else QLineEdit.Password
            )
        )
        qd_key_row.addWidget(self._qd_api_key)
        qd_key_row.addWidget(_qd_show_btn)
        qd_conn_f.addRow(_t("API Key:"), qd_key_row)

        self._qd_prefer_grpc = QCheckBox(_t("使用 gRPC（高性能，需要 gRPC 端口，通常为 6334）"))
        self._qd_prefer_grpc.setChecked(True)
        qd_conn_f.addRow("", self._qd_prefer_grpc)

        qd_outer.addWidget(qd_conn_gb)

        # ── BGE-M3 嵌入服务 ────────────────────────────────────────
        bge_gb = QGroupBox(_t("BGE-M3 嵌入服务"))
        bge_f = QFormLayout(bge_gb)
        bge_f.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        bge_f.setSpacing(5)

        self._bge_url = QLineEdit()
        self._bge_url.setPlaceholderText(_t("如 http://127.0.0.1:7997"))
        bge_f.addRow(_t("服务地址（Base URL）:"), self._bge_url)

        bge_key_row = QHBoxLayout()
        self._bge_api_key = QLineEdit()
        self._bge_api_key.setEchoMode(QLineEdit.Password)
        self._bge_api_key.setPlaceholderText(_t("API Key（可留空）"))
        _bge_show_btn = QPushButton("👁")
        _bge_show_btn.setFixedWidth(32)
        _bge_show_btn.setCheckable(True)
        _bge_show_btn.toggled.connect(
            lambda c: self._bge_api_key.setEchoMode(
                QLineEdit.Normal if c else QLineEdit.Password
            )
        )
        bge_key_row.addWidget(self._bge_api_key)
        bge_key_row.addWidget(_bge_show_btn)
        bge_f.addRow(_t("API Key:"), bge_key_row)

        self._bge_model = QLineEdit()
        self._bge_model.setText("BAAI/bge-m3")
        bge_f.addRow(_t("模型名称:"), self._bge_model)

        self._bge_timeout = QSpinBox()
        self._bge_timeout.setRange(5, 600)
        self._bge_timeout.setValue(120)
        self._bge_timeout.setSuffix(_t(" 秒"))
        bge_f.addRow(_t("超时:"), self._bge_timeout)

        qd_outer.addWidget(bge_gb)

        # ── 查询行为 ───────────────────────────────────────────────
        qd_q_gb = QGroupBox(_t("查询行为"))
        qd_q_f = QFormLayout(qd_q_gb)
        qd_q_f.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        qd_q_f.setSpacing(5)

        self._qd_mode = QComboBox()
        self._qd_mode.addItem(_t("hybrid（稠密+稀疏 RRF 融合，推荐）"), "hybrid")
        self._qd_mode.addItem(_t("dense（仅稠密向量）"),                "dense")
        self._qd_mode.addItem(_t("sparse（仅稀疏向量）"),               "sparse")
        qd_q_f.addRow(_t("查询模式:"), self._qd_mode)

        self._qd_threshold = QDoubleSpinBox()
        self._qd_threshold.setRange(0.0, 1.0)
        self._qd_threshold.setSingleStep(0.05)
        self._qd_threshold.setDecimals(2)
        self._qd_threshold.setValue(0.0)
        self._qd_threshold.setToolTip(_t("0.0 = 不过滤；设置后只返回相似度 ≥ 阈值的结果"))
        qd_q_f.addRow(_t("相似度阈值:"), self._qd_threshold)

        self._qd_limit = QSpinBox()
        self._qd_limit.setRange(1, 500)
        self._qd_limit.setValue(20)
        qd_q_f.addRow(_t("默认返回数量:"), self._qd_limit)

        self._qd_timeout = QSpinBox()
        self._qd_timeout.setRange(5, 300)
        self._qd_timeout.setValue(30)
        self._qd_timeout.setSuffix(_t(" 秒"))
        qd_q_f.addRow(_t("查询超时:"), self._qd_timeout)

        qd_outer.addWidget(qd_q_gb)


        # ── 返回字段配置 ───────────────────────────────────────────
        qd_fld_gb = QGroupBox(_t("返回字段配置（空=返回全部字段）"))
        qd_fld_gb.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        qd_fld_v = QVBoxLayout(qd_fld_gb)
        qd_fld_v.setSpacing(4)

        fld_note = QLabel(
            _t("点击「🔍 自动获取字段」从 Qdrant collection 读取所有 payload 字段。\n"
            "✅ 勾选的字段将被返回；无任何勾选则返回全部字段。")
        )
        fld_note.setStyleSheet("color:#666; font-size:11px;")
        qd_fld_v.addWidget(fld_note)

        self._qd_field_list = QListWidget()
        self._qd_field_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._qd_field_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self._qd_field_list.setToolTip(_t("勾选字段将被返回；无勾选 = 返回所有字段"))
        qd_fld_v.addWidget(self._qd_field_list)

        self._qd_field_status = QLabel(_t("（尚未获取字段，请先测试连接或点击下方按钮）"))
        self._qd_field_status.setStyleSheet("color:#888; font-size:11px;")
        qd_fld_v.addWidget(self._qd_field_status)

        fld_btn_row = QHBoxLayout()
        self._qd_fetch_fields_btn = QPushButton(_t("🔍 自动获取字段"))
        self._qd_fetch_fields_btn.clicked.connect(self._on_qd_fetch_fields)
        _qd_check_all_btn = QPushButton(_t("全选"))
        _qd_check_all_btn.setFixedWidth(60)
        _qd_check_all_btn.clicked.connect(self._qd_field_check_all)
        _qd_uncheck_all_btn = QPushButton(_t("全取消"))
        _qd_uncheck_all_btn.setFixedWidth(60)
        _qd_uncheck_all_btn.clicked.connect(self._qd_field_uncheck_all)
        _qd_remove_btn = QPushButton(_t("🗑 移除"))
        _qd_remove_btn.setFixedWidth(65)
        _qd_remove_btn.setToolTip(_t("移除列表中当前高亮的字段"))
        _qd_remove_btn.clicked.connect(self._qd_field_remove_selected)
        fld_btn_row.addWidget(self._qd_fetch_fields_btn)
        fld_btn_row.addWidget(_qd_check_all_btn)
        fld_btn_row.addWidget(_qd_uncheck_all_btn)
        fld_btn_row.addWidget(_qd_remove_btn)
        fld_btn_row.addStretch()
        qd_fld_v.addLayout(fld_btn_row)

        qd_outer.addWidget(qd_fld_gb)

        # ── 前置过滤字段 ────────────────────────────────────────────
        qd_filter_gb = QGroupBox(_t("前置过滤字段（静态 Payload Filter，逻辑 AND）"))
        qd_filter_v = QVBoxLayout(qd_filter_gb)
        qd_filter_v.setSpacing(4)

        filter_note = QLabel(
            _t("每行为一个过滤条件，查询时作为 Qdrant must filter 生效。\n"
            "字段名支持后缀：_min/_max（数值范围）。示例：material_status = 有效")
        )
        filter_note.setStyleSheet("color:#666; font-size:11px;")
        qd_filter_v.addWidget(filter_note)

        self._qd_filter_table = QTableWidget(0, 2)
        self._qd_filter_table.setHorizontalHeaderLabels([_t("过滤字段"), _t("匹配值")])
        _qf_hdr = self._qd_filter_table.horizontalHeader()
        _qf_hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        _qf_hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        self._qd_filter_table.setMaximumHeight(120)
        self._qd_filter_table.setAlternatingRowColors(True)
        qd_filter_v.addWidget(self._qd_filter_table)

        qf_btn_row = QHBoxLayout()
        _qf_add = QPushButton(_t("➕ 添加"))
        _qf_add.setFixedWidth(70)
        _qf_add.clicked.connect(lambda: self._qd_filter_add_row())
        _qf_del = QPushButton(_t("🗑 删除"))
        _qf_del.setFixedWidth(70)
        _qf_del.clicked.connect(self._qd_filter_del_row)
        qf_btn_row.addWidget(_qf_add)
        qf_btn_row.addWidget(_qf_del)
        qf_btn_row.addStretch()
        qd_filter_v.addLayout(qf_btn_row)

        qd_outer.addWidget(qd_filter_gb)
        fl.addWidget(self._qdrant_gb)

        # ── sql_table 专项参数 ────────────────────────────────────
        self._sql_gb = QGroupBox(_t("SQL 数据库参数（sql_table）"))
        self._sql_gb.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        sql_outer = QVBoxLayout(self._sql_gb)
        sql_outer.setSpacing(8)

        # 连接参数
        sql_conn_gb = QGroupBox(_t("数据库连接"))
        sql_conn_f  = QFormLayout(sql_conn_gb)
        sql_conn_f.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)
        sql_conn_f.setSpacing(5)

        self._sql_db_type = QComboBox()
        self._sql_db_type.addItem(_t("PostgreSQL"), "postgresql")
        self._sql_db_type.addItem(_t("SQLite（本地文件）"), "sqlite")
        self._sql_db_type.currentIndexChanged.connect(self._sql_on_dbtype_changed)
        sql_conn_f.addRow(_t("数据库类型:"), self._sql_db_type)

        # PostgreSQL 参数行（随类型切换显隐）
        self._sql_pg_host = QLineEdit()
        self._sql_pg_host.setPlaceholderText(_t("IP 地址或域名，如 127.0.0.1"))
        sql_conn_f.addRow(_t("主机:"), self._sql_pg_host)

        self._sql_pg_port = QSpinBox()
        self._sql_pg_port.setRange(1, 65535)
        self._sql_pg_port.setValue(5432)
        sql_conn_f.addRow(_t("端口:"), self._sql_pg_port)

        self._sql_pg_dbname = QLineEdit()
        self._sql_pg_dbname.setPlaceholderText(_t("数据库名称"))
        sql_conn_f.addRow(_t("数据库名:"), self._sql_pg_dbname)

        self._sql_pg_user = QLineEdit()
        self._sql_pg_user.setPlaceholderText(_t("用户名"))
        sql_conn_f.addRow(_t("用户名:"), self._sql_pg_user)

        sql_pass_row = QHBoxLayout()
        self._sql_pg_pass = QLineEdit()
        self._sql_pg_pass.setEchoMode(QLineEdit.Password)
        self._sql_pg_pass.setPlaceholderText(_t("密码"))
        _sql_show_btn = QPushButton("👁")
        _sql_show_btn.setFixedWidth(32)
        _sql_show_btn.setCheckable(True)
        _sql_show_btn.toggled.connect(
            lambda c: self._sql_pg_pass.setEchoMode(
                QLineEdit.Normal if c else QLineEdit.Password
            )
        )
        sql_pass_row.addWidget(self._sql_pg_pass)
        sql_pass_row.addWidget(_sql_show_btn)
        sql_conn_f.addRow(_t("密码:"), sql_pass_row)

        # SQLite 参数行
        sql_path_row = QHBoxLayout()
        self._sql_db_path = QLineEdit()
        self._sql_db_path.setPlaceholderText("SQLite 数据库文件路径，如 C:/data/db.sqlite")
        _sql_browse_btn = QPushButton("📂")
        _sql_browse_btn.setFixedWidth(36)
        _sql_browse_btn.setToolTip(_t("浏览 SQLite 文件"))
        _sql_browse_btn.clicked.connect(self._sql_browse_file)
        sql_path_row.addWidget(self._sql_db_path)
        sql_path_row.addWidget(_sql_browse_btn)
        sql_conn_f.addRow(_t("数据库文件:"), sql_path_row)

        self._sql_timeout = QSpinBox()
        self._sql_timeout.setRange(5, 120)
        self._sql_timeout.setValue(15)
        self._sql_timeout.setSuffix(_t(" 秒"))
        sql_conn_f.addRow(_t("超时:"), self._sql_timeout)

        sql_outer.addWidget(sql_conn_gb)
        self._sql_pg_widgets = [self._sql_pg_host, self._sql_pg_port,
                                 self._sql_pg_dbname, self._sql_pg_user,
                                 self._sql_pg_pass]
        self._sql_sqlite_widgets = [self._sql_db_path]

        # 表选择
        tbl_sel_gb = QGroupBox(_t("目标表 / 视图"))
        tbl_sel_f  = QHBoxLayout(tbl_sel_gb)
        self._sql_table_combo = QComboBox()
        self._sql_table_combo.setEditable(True)
        self._sql_table_combo.setPlaceholderText(_t("请先测试连接获取表列表"))
        self._sql_table_combo.currentTextChanged.connect(self._sql_on_table_changed)
        self._sql_fetch_tables_btn = QPushButton(_t("🔍 获取表列表"))
        self._sql_fetch_tables_btn.clicked.connect(self._sql_fetch_tables)
        tbl_sel_f.addWidget(self._sql_table_combo, 1)
        tbl_sel_f.addWidget(self._sql_fetch_tables_btn)
        sql_outer.addWidget(tbl_sel_gb)

        # 字段配置
        fld_cfg_gb = QGroupBox(_t("字段配置"))
        fld_cfg_gb.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        fld_cfg_v  = QVBoxLayout(fld_cfg_gb)
        fld_cfg_v.setSpacing(4)

        sql_fld_note = QLabel(
            _t("✅ 返回字段：勾选后该字段将出现在查询结果中\n"
            "🔍 查询字段：勾选后可作为 WHERE 搜索条件（对应映射组输入映射）\n"
            "点击「🔍 获取字段」从数据库自动读取表结构")
        )
        sql_fld_note.setStyleSheet("color:#666; font-size:11px;")
        fld_cfg_v.addWidget(sql_fld_note)

        self._sql_field_table = QTableWidget(0, 3)
        self._sql_field_table.setHorizontalHeaderLabels([_t("字段名"), _t("返回"), _t("查询")])
        _sfh = self._sql_field_table.horizontalHeader()
        _sfh.setSectionResizeMode(0, QHeaderView.Stretch)
        _sfh.setSectionResizeMode(1, QHeaderView.Fixed)
        _sfh.setSectionResizeMode(2, QHeaderView.Fixed)
        self._sql_field_table.setColumnWidth(1, 50)
        self._sql_field_table.setColumnWidth(2, 50)
        self._sql_field_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._sql_field_table.setAlternatingRowColors(True)
        fld_cfg_v.addWidget(self._sql_field_table)

        sql_fld_btn_row = QHBoxLayout()
        self._sql_fetch_fields_btn = QPushButton(_t("🔍 获取字段"))
        self._sql_fetch_fields_btn.clicked.connect(self._sql_fetch_fields)
        _sql_chk_all_ret = QPushButton(_t("全选返回"))
        _sql_chk_all_ret.setFixedWidth(72)
        _sql_chk_all_ret.clicked.connect(lambda: self._sql_field_check_col(1, True))
        _sql_unchk_all_ret = QPushButton(_t("取消返回"))
        _sql_unchk_all_ret.setFixedWidth(72)
        _sql_unchk_all_ret.clicked.connect(lambda: self._sql_field_check_col(1, False))
        _sql_chk_all_srch = QPushButton(_t("全选查询"))
        _sql_chk_all_srch.setFixedWidth(72)
        _sql_chk_all_srch.clicked.connect(lambda: self._sql_field_check_col(2, True))
        _sql_unchk_all_srch = QPushButton(_t("取消查询"))
        _sql_unchk_all_srch.setFixedWidth(72)
        _sql_unchk_all_srch.clicked.connect(lambda: self._sql_field_check_col(2, False))
        sql_fld_btn_row.addWidget(self._sql_fetch_fields_btn)
        sql_fld_btn_row.addSpacing(8)
        sql_fld_btn_row.addWidget(_sql_chk_all_ret)
        sql_fld_btn_row.addWidget(_sql_unchk_all_ret)
        sql_fld_btn_row.addSpacing(8)
        sql_fld_btn_row.addWidget(_sql_chk_all_srch)
        sql_fld_btn_row.addWidget(_sql_unchk_all_srch)
        sql_fld_btn_row.addStretch()
        fld_cfg_v.addLayout(sql_fld_btn_row)
        sql_outer.addWidget(fld_cfg_gb)

        # 查询匹配模式
        match_gb = QGroupBox(_t("查询匹配模式"))
        match_f  = QFormLayout(match_gb)
        match_f.setSpacing(5)
        self._sql_match_mode = QComboBox()
        self._sql_match_mode.addItem(_t("模糊匹配（ILIKE / LIKE，推荐）"), "ilike")
        self._sql_match_mode.addItem(_t("精确匹配（= ）"), "exact")
        match_f.addRow(_t("匹配方式:"), self._sql_match_mode)
        sql_outer.addWidget(match_gb)

        fl.addWidget(self._sql_gb)

        fl.addStretch()

        # ── 底部操作按钮 ─────────────────────────────────────────
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #e0e0e0;")
        fl.addWidget(sep)

        action_row = QHBoxLayout()
        action_row.setSpacing(8)

        self._test_btn = QPushButton(_t("🔌 测试连接"))
        self._test_btn.setFixedWidth(110)
        self._test_btn.setToolTip(_t("使用当前表单中的参数实例化后端并测试连通性"))
        self._test_btn.clicked.connect(self._on_test)

        self._test_label = QLabel("")
        self._test_label.setStyleSheet("font-size: 11px;")

        self._apply_btn = QPushButton(_t("✔ 应用修改"))
        self._apply_btn.setFixedWidth(110)
        self._apply_btn.setStyleSheet(
            "background:#0078d4; color:white; border-radius:4px; padding:5px;"
        )
        self._apply_btn.clicked.connect(self._on_apply)

        action_row.addWidget(self._test_btn)
        action_row.addWidget(self._test_label, 1)
        action_row.addWidget(self._apply_btn)
        fl.addLayout(action_row)

    # ── 列表刷新 ─────────────────────────────────────────────────

    def _refresh_type_combo(self) -> None:
        """根据当前已注册的后端插件动态更新类型下拉框"""
        from core.query_base import BACKEND_REGISTRY
        cur_data = self._type_combo.currentData()
        self._type_combo.blockSignals(True)
        self._type_combo.clear()
        for t in sorted(BACKEND_REGISTRY.keys()):
            self._type_combo.addItem(t, t)
        # 兜底选项：自定义 / 未来新后端
        self._type_combo.addItem("自定义 (custom)", "custom")
        # 恢复之前的选中
        idx = self._type_combo.findData(cur_data)
        if idx >= 0:
            self._type_combo.setCurrentIndex(idx)
        self._type_combo.blockSignals(False)

    def refresh(self) -> None:
        """每次标签页切到此 Tab 时调用，刷新模块列表"""
        self._refresh_type_combo()
        self._list.blockSignals(True)
        self._list.clear()

        for mod_id, mod_cfg in self._app.config_mgr.get_query_modules().items():
            enabled = mod_cfg.get("enabled", True)
            btype = mod_cfg.get("type", "?")
            mod_name = mod_cfg.get("name", "").strip()
            status_text = _t("● 已启用") if enabled else _t("○ 已暂停")
            display_name = mod_name if mod_name else mod_id
            item = QListWidgetItem(
                f"{'⚡' if enabled else '⊘'} {display_name}\n"
                f"   {mod_id}  [{btype}]  {status_text}"
            )
            item.setData(Qt.UserRole, mod_id)
            if not enabled:
                item.setForeground(QColor("#9e9e9e"))
            self._list.addItem(item)

        self._list.blockSignals(False)
        self._placeholder.show()
        self._form_scroll.hide()
        self._toggle_btn.setEnabled(False)
        self._current_mod_id = None

    # ── 事件处理 ─────────────────────────────────────────────────

    def _on_select(self, row: int) -> None:
        if row < 0:
            return
        mod_id = self._list.item(row).data(Qt.UserRole)
        cfg = self._app.config_mgr.get_query_modules().get(mod_id, {})
        self._load_into_form(mod_id, cfg, is_new=False)

    def _load_into_form(self, mod_id: str, cfg: dict, *, is_new: bool) -> None:
        """将模块配置填入右侧表单，切换占位符和表单的可见性"""
        self._current_mod_id = mod_id
        self._is_new = is_new

        self._placeholder.hide()
        self._form_scroll.show()

        # 基本信息
        self._id_edit.setText(mod_id)
        self._name_edit.setText(cfg.get("name", ""))
        self._id_edit.setReadOnly(True)
        self._id_edit.setStyleSheet("background: #fafafa; color: #555;")

        self._refresh_type_combo()
        b_type = cfg.get("type", "vector_api")
        idx = self._type_combo.findData(b_type)
        if idx < 0:
            # 未知类型：在最前面动态插入，并使用自定义键值对兜底
            self._type_combo.insertItem(0, f"{b_type}  （未知类型）", b_type)
            idx = 0
        self._type_combo.setCurrentIndex(idx)

        self._enabled_cb.setChecked(cfg.get("enabled", True))

        # vector_api 参数
        self._v_url.setText(cfg.get("base_url", ""))
        self._v_timeout.setValue(cfg.get("timeout", 30))

        # pg 参数
        db = cfg.get("database", {})
        self._pg_host.setText(db.get("host", ""))
        self._pg_port.setValue(int(db.get("port", 5432)))
        self._pg_dbname.setText(db.get("dbname", ""))
        self._pg_user.setText(db.get("user", ""))
        self._pg_pass.setText(db.get("password", ""))
        self._show_pass_btn.setChecked(False)
        self._pg_timeout.setValue(cfg.get("timeout", 15))

        # excel_file 参数
        self._excel_src_list.clear()
        for src in cfg.get("sources", []):
            src_type = src.get("type", "file")
            path     = src.get("path", "")
            if src_type == "file":
                label = f"📄 {path}"
            else:
                recursive = src.get("recursive", False)
                label = f"📁 {path}{'  （含子目录）' if recursive else ''}"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, src)
            self._excel_src_list.addItem(item)
        self._excel_has_header.setChecked(bool(cfg.get("has_header", True)))
        
        # 解析 sheet 配置并勾选
        raw_sheet = cfg.get("sheet", "0")
        target_sheets = []
        if isinstance(raw_sheet, (list, tuple)):
            target_sheets = [str(x) for x in raw_sheet]
        elif isinstance(raw_sheet, (int, float)):
            target_sheets = [str(int(raw_sheet))]
        elif isinstance(raw_sheet, str):
            if "," in raw_sheet or "，" in raw_sheet:
                target_sheets = [s.strip() for s in raw_sheet.replace("，", ",").split(",") if s.strip()]
            else:
                target_sheets = [raw_sheet.strip()]
        else:
            target_sheets = ["0"]

        self._excel_sheet_list.clear()
        loaded_sheets = []
        first_filepath = ""
        for i in range(self._excel_src_list.count()):
            src_item = self._excel_src_list.item(i)
            src_data = src_item.data(Qt.UserRole)
            if src_data and src_data.get("type") == "file":
                first_filepath = src_data.get("path", "")
                break

        import os
        if first_filepath and os.path.exists(first_filepath):
            try:
                # 优先用 calamine
                try:
                    from python_calamine import CalamineWorkbook
                    wb = CalamineWorkbook.from_path(first_filepath)
                    loaded_sheets = wb.sheet_names
                except ImportError:
                    import openpyxl
                    wb = openpyxl.load_workbook(first_filepath, read_only=True)
                    loaded_sheets = wb.sheetnames
                    wb.close()
            except Exception as e:
                logger.error("自动加载工作表失败: %s", e)

        if loaded_sheets:
            for idx, s_name in enumerate(loaded_sheets):
                item = QListWidgetItem(s_name)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                is_checked = (s_name in target_sheets or str(idx) in target_sheets or (not target_sheets and idx == 0))
                item.setCheckState(Qt.Checked if is_checked else Qt.Unchecked)
                self._excel_sheet_list.addItem(item)
        else:
            for s_val in target_sheets:
                item = QListWidgetItem(s_val)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Checked)
                self._excel_sheet_list.addItem(item)

        self._excel_max_rows.setValue(int(cfg.get("max_rows", 100_000)))
        self._excel_col_table.setRowCount(0)
        for col_cfg in cfg.get("columns", []):
            self._excel_add_col_row(
                str(col_cfg.get("col", "")),
                str(col_cfg.get("field", "")),
                checked=True
            )
        self._adjust_excel_col_table_height()

        # qdrant_vector 参数
        self._qd_host.setText(cfg.get("qdrant_host", ""))
        self._qd_port.setValue(int(cfg.get("qdrant_port", 6334)))
        self._qd_collection.setText(cfg.get("qdrant_collection", ""))
        self._qd_api_key.setText(cfg.get("qdrant_api_key", ""))
        self._qd_prefer_grpc.setChecked(bool(cfg.get("qdrant_prefer_grpc", True)))
        self._bge_url.setText(cfg.get("embed_base_url", ""))
        self._bge_api_key.setText(cfg.get("embed_api_key", ""))
        self._bge_model.setText(cfg.get("embed_model", "BAAI/bge-m3"))
        self._bge_timeout.setValue(int(cfg.get("embed_timeout", 120)))
        mode_idx = self._qd_mode.findData(cfg.get("search_mode", "hybrid"))
        self._qd_mode.setCurrentIndex(mode_idx if mode_idx >= 0 else 0)
        self._qd_threshold.setValue(float(cfg.get("score_threshold", 0.0)))
        self._qd_limit.setValue(int(cfg.get("limit", 20)))
        self._qd_timeout.setValue(int(cfg.get("timeout", 30)))
        # 加载已保存的返回字段列表（带勾选状态）
        saved_fields = cfg.get("return_fields") or []
        self._qd_field_list.clear()
        self._qd_field_status.setText("（已加载保存的字段配置）" if saved_fields else "（尚未获取字段，请先测试连接或点击自动获取）")
        for f in saved_fields:
            item = QListWidgetItem(f)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)
            self._qd_field_list.addItem(item)
        self._adjust_qd_field_list_height()
        # 加载前置过滤
        self._qd_filter_table.setRowCount(0)
        for fkey, fval in (cfg.get("default_filters") or {}).items():
            self._qd_filter_add_row(fkey, str(fval))

        # 通用键值对：从 cfg 中剔除已由专项表单覆盖的标准键
        self._kv_table.setRowCount(0)
        _std_keys = {
            "type", "enabled", "base_url", "timeout", "database",
            "sources", "has_header", "sheet", "max_rows", "columns",
            "qdrant_host", "qdrant_port", "qdrant_collection", "qdrant_api_key",
            "qdrant_prefer_grpc", "embed_base_url", "embed_api_key", "embed_model",
            "embed_timeout", "search_mode", "score_threshold", "limit",
            "return_fields", "default_filters",
        }
        for k, v in cfg.items():
            if k not in _std_keys:
                self._kv_add_row(k, str(v))

        # sql_table 参数
        db_type_idx = self._sql_db_type.findData(cfg.get("db_type", "postgresql"))
        self._sql_db_type.setCurrentIndex(db_type_idx if db_type_idx >= 0 else 0)
        self._sql_pg_host.setText(cfg.get("host", ""))
        self._sql_pg_port.setValue(int(cfg.get("port", 5432)))
        self._sql_pg_dbname.setText(cfg.get("dbname", ""))
        self._sql_pg_user.setText(cfg.get("user", ""))
        self._sql_pg_pass.setText(cfg.get("password", ""))
        self._sql_db_path.setText(cfg.get("db_path", ""))
        self._sql_timeout.setValue(int(cfg.get("timeout", 15)))
        match_idx = self._sql_match_mode.findData(cfg.get("match_mode", "ilike"))
        self._sql_match_mode.setCurrentIndex(match_idx if match_idx >= 0 else 0)
        # 字段列表：从已保存的 return_fields / search_fields 恢复
        self._sql_field_table.setRowCount(0)
        saved_ret   = set(cfg.get("return_fields",  []) or [])
        saved_srch  = set(cfg.get("search_fields", []) or [])
        for f in sorted(saved_ret | saved_srch):
            self._sql_add_field_row(f, f in saved_ret, f in saved_srch)
        self._adjust_sql_field_table_height()
        # 表下拉恢复
        saved_table = cfg.get("table", "")
        self._sql_table_combo.blockSignals(True)
        self._sql_table_combo.clear()
        if saved_table:
            self._sql_table_combo.addItem(saved_table)
            self._sql_table_combo.setCurrentText(saved_table)
        self._sql_table_combo.blockSignals(False)
        self._sql_on_dbtype_changed()  # 初始化显隐状态


        # 动态切换专项参数区域
        self._on_type_changed()

        # 暂停/启用按钮
        self._toggle_btn.setEnabled(not is_new)
        enabled = cfg.get("enabled", True)
        self._toggle_btn.setText("⏸ 暂停" if enabled else "▶ 启用")
        self._toggle_btn.setToolTip("暂停此模块" if enabled else "启用此模块")

        self._test_label.setText("")

    def _on_type_changed(self) -> None:
        """根据选中的后端类型动态显示专项参数区域"""
        b_type = self._type_combo.currentData() or ""
        self._vector_gb.setVisible(b_type in self._URL_TYPES)
        self._pg_gb.setVisible(b_type in ("pg_plan", "pg_spare_parts"))
        self._excel_gb.setVisible(b_type == "excel_file")
        self._qdrant_gb.setVisible(b_type == "qdrant_vector")
        self._sql_gb.setVisible(b_type == "sql_table")
        # 不在任何专项类型中，则显示通用键值对兜底
        self._custom_gb.setVisible(b_type not in self._SPECIALIZED_TYPES)
        if b_type == "excel_file":
            self._adjust_excel_col_table_height()
        elif b_type == "sql_table":
            self._adjust_sql_field_table_height()
        elif b_type == "qdrant_vector":
            self._adjust_qd_field_list_height()

    def _on_add(self) -> None:
        mod_id = ConfigManager.generate_id("qm_")
        # 确保生成的 ID 不冲突
        while mod_id in self._app.config_mgr.get_query_modules():
            mod_id = ConfigManager.generate_id("qm_")
        # 用空配置打开新建表单
        self._load_into_form(mod_id, {}, is_new=True)

    def _on_delete(self) -> None:
        row = self._list.currentRow()
        if row < 0:
            QMessageBox.information(self, _t("提示"), _t("请先在左侧选中要删除的模块"))
            return
        mod_id = self._list.item(row).data(Qt.UserRole)
        refs = self._app.config_mgr.get_referencing_mapping_groups(mod_id)
        if refs:
            ref_list = "\n".join(f"  • {r}" for r in refs)
            ans = QMessageBox.warning(
                self, _t("确认删除（存在引用）"),
                _t("以下映射组引用了模块「{mod_id}」：\n{ref_list}\n\n"
                   "删除后这些映射组的查询模块字段将被清空，确认删除？").format(mod_id=mod_id, ref_list=ref_list),
                QMessageBox.Yes | QMessageBox.No,
            )
        else:
            ans = QMessageBox.question(
                self, _t("确认删除"),
                _t("确定删除查询模块「{mod_id}」？").format(mod_id=mod_id)
            )
        if ans == QMessageBox.Yes:
            self._app.config_mgr.delete_query_module(mod_id)
            self._app.query_modules.pop(mod_id, None)
            self.refresh()

    def _on_toggle(self) -> None:
        """暂停/启用当前选中模块，立即生效"""
        if not self._current_mod_id:
            return
        all_cfgs = self._app.config_mgr.get_query_modules()
        cfg_dict = dict(all_cfgs.get(self._current_mod_id, {}))
        new_enabled = not cfg_dict.get("enabled", True)
        cfg_dict["enabled"] = new_enabled

        # 持久化
        self._app.config_mgr.save_query_module(self._current_mod_id, cfg_dict)

        # 即时内存生效
        if new_enabled:
            self._app._reinit_single_module(self._current_mod_id, cfg_dict)
        else:
            self._app.query_modules.pop(self._current_mod_id, None)

        # 更新按钮文字 & 勾选框
        self._enabled_cb.setChecked(new_enabled)
        self._toggle_btn.setText("⏸ 暂停" if new_enabled else "▶ 启用")
        self._toggle_btn.setToolTip("暂停此模块" if new_enabled else "启用此模块")

        # 刷新列表中的那一行
        self._update_list_item(self._current_mod_id, cfg_dict)

    def _update_list_item(self, mod_id: str, cfg: dict) -> None:
        """只更新列表中的指定行，不重绘整个列表"""
        for i in range(self._list.count()):
            if self._list.item(i).data(Qt.UserRole) == mod_id:
                item = self._list.item(i)
                enabled = cfg.get("enabled", True)
                btype = cfg.get("type", "?")
                mod_name = cfg.get("name", "").strip()
                status_text = _t("● 已启用") if enabled else _t("○ 已暂停")
                display_name = mod_name if mod_name else mod_id
                item.setText(
                    f"{'⚡' if enabled else '⊘'} {display_name}\n"
                    f"   {mod_id}  [{btype}]  {status_text}"
                )
                item.setForeground(
                    QColor("#212121") if enabled else QColor("#9e9e9e")
                )
                break

    def _on_apply(self) -> None:
        """验证并保存表单，即时重载模块"""
        mod_id = self._id_edit.text().strip()
        if not mod_id:
            QMessageBox.warning(self, _t("错误"), _t("模块 ID 不能为空"))
            return
        if self._is_new and mod_id in self._app.config_mgr.get_query_modules():
            QMessageBox.warning(self, _t("错误"), _t("模块 ID「{mod_id}」已存在，请换一个").format(mod_id=mod_id))
            return

        b_type = self._type_combo.currentData() or "custom"
        enabled = self._enabled_cb.isChecked()
        mod_name = self._name_edit.text().strip()

        new_cfg: dict[str, Any] = {"type": b_type, "enabled": enabled}
        if mod_name:
            new_cfg["name"] = mod_name

        if b_type in self._URL_TYPES:
            url = self._v_url.text().strip()
            if not url:
                QMessageBox.warning(self, _t("错误"), _t("API 地址不能为空"))
                return
            new_cfg["base_url"] = url
            new_cfg["timeout"] = self._v_timeout.value()

        elif b_type in ("pg_plan", "pg_spare_parts"):
            host = self._pg_host.text().strip()
            if not host:
                QMessageBox.warning(self, _t("错误"), _t("数据库主机不能为空"))
                return
            new_cfg["timeout"] = self._pg_timeout.value()
            new_cfg["database"] = {
                "host":     host,
                "port":     self._pg_port.value(),
                "dbname":   self._pg_dbname.text().strip(),
                "user":     self._pg_user.text().strip(),
                "password": self._pg_pass.text(),
            }

        elif b_type == "excel_file":
            sources = [
                self._excel_src_list.item(i).data(Qt.UserRole)
                for i in range(self._excel_src_list.count())
                if self._excel_src_list.item(i).data(Qt.UserRole)
            ]
            if not sources:
                QMessageBox.warning(self, _t("错误"), _t("请至少添加一个数据源（文件或目录）"))
                return
            new_cfg["sources"] = sources
            new_cfg["has_header"] = self._excel_has_header.isChecked()
            checked_sheets = []
            for i in range(self._excel_sheet_list.count()):
                item = self._excel_sheet_list.item(i)
                if item.checkState() == Qt.Checked:
                    checked_sheets.append(item.text())
            sheet_val = ",".join(checked_sheets) if checked_sheets else "0"
            new_cfg["sheet"] = sheet_val
            new_cfg["max_rows"] = self._excel_max_rows.value()
            columns = []
            for r in range(self._excel_col_table.rowCount()):
                widget = self._excel_col_table.cellWidget(r, 0)
                is_checked = True
                if widget:
                    cb = widget.findChild(QCheckBox)
                    if cb:
                        is_checked = cb.isChecked()
                if not is_checked:
                    continue
                ci = self._excel_col_table.item(r, 1)
                fi = self._excel_col_table.item(r, 2)
                col_val   = ci.text().strip() if ci else ""
                field_val = fi.text().strip() if fi else ""
                if col_val:
                    columns.append({"col": col_val, "field": field_val or col_val})
            new_cfg["columns"] = columns

        elif b_type == "qdrant_vector":
            qd_host = self._qd_host.text().strip()
            qd_coll = self._qd_collection.text().strip()
            bge_url = self._bge_url.text().strip()
            if not qd_host:
                QMessageBox.warning(self, _t("错误"), _t("Qdrant 主机不能为空"))
                return
            if not qd_coll:
                QMessageBox.warning(self, _t("错误"), _t("Qdrant 数据集（Collection）不能为空"))
                return
            if not bge_url:
                QMessageBox.warning(self, _t("错误"), _t("BGE-M3 服务地址不能为空"))
                return
            # 收集勾选的返回字段
            return_fields = [
                self._qd_field_list.item(i).text()
                for i in range(self._qd_field_list.count())
                if self._qd_field_list.item(i).checkState() == Qt.Checked
            ]
            # 收集前置过滤
            default_filters: dict = {}
            for r in range(self._qd_filter_table.rowCount()):
                ki = self._qd_filter_table.item(r, 0)
                vi = self._qd_filter_table.item(r, 1)
                if ki and ki.text().strip():
                    default_filters[ki.text().strip()] = vi.text().strip() if vi else ""
            new_cfg.update({
                "qdrant_host":         qd_host,
                "qdrant_port":         self._qd_port.value(),
                "qdrant_collection":   qd_coll,
                "qdrant_api_key":      self._qd_api_key.text(),
                "qdrant_prefer_grpc":  self._qd_prefer_grpc.isChecked(),
                "embed_base_url":      bge_url,
                "embed_api_key":       self._bge_api_key.text(),
                "embed_model":         self._bge_model.text().strip() or "BAAI/bge-m3",
                "embed_timeout":       self._bge_timeout.value(),
                "search_mode":         self._qd_mode.currentData() or "hybrid",
                "score_threshold":     self._qd_threshold.value(),
                "limit":               self._qd_limit.value(),
                "timeout":             self._qd_timeout.value(),
                "return_fields":       return_fields,
                "default_filters":     default_filters,
            })

        elif b_type == "sql_table":
            db_type = self._sql_db_type.currentData() or "postgresql"
            table   = self._sql_table_combo.currentText().strip()
            if not table:
                QMessageBox.warning(self, _t("错误"), _t("请选择或输入目标表/视图名称"))
                return
            if db_type == "postgresql" and not self._sql_pg_host.text().strip():
                QMessageBox.warning(self, _t("错误"), _t("请填写数据库主机地址"))
                return
            if db_type == "sqlite" and not self._sql_db_path.text().strip():
                QMessageBox.warning(self, _t("错误"), _t("请填写 SQLite 数据库文件路径"))
                return
            ret_fields:  list[str] = []
            srch_fields: list[str] = []
            for r in range(self._sql_field_table.rowCount()):
                fn_item = self._sql_field_table.item(r, 0)
                cb_ret  = self._sql_field_table.cellWidget(r, 1)
                cb_srch = self._sql_field_table.cellWidget(r, 2)
                fn = fn_item.text().strip() if fn_item else ""
                if not fn:
                    continue
                if cb_ret  and cb_ret.isChecked():
                    ret_fields.append(fn)
                if cb_srch and cb_srch.isChecked():
                    srch_fields.append(fn)
            new_cfg.update({
                "db_type":       db_type,
                "table":         table,
                "host":          self._sql_pg_host.text().strip(),
                "port":          self._sql_pg_port.value(),
                "dbname":        self._sql_pg_dbname.text().strip(),
                "user":          self._sql_pg_user.text().strip(),
                "password":      self._sql_pg_pass.text(),
                "db_path":       self._sql_db_path.text().strip(),
                "timeout":       self._sql_timeout.value(),
                "match_mode":    self._sql_match_mode.currentData() or "ilike",
                "return_fields":  ret_fields,
                "search_fields": srch_fields,
            })

        else:
            # 自定义 / 未知类型：从键值对表格收集所有参数
            for r in range(self._kv_table.rowCount()):
                ki = self._kv_table.item(r, 0)
                vi = self._kv_table.item(r, 1)
                if ki and ki.text().strip():
                    new_cfg[ki.text().strip()] = vi.text().strip() if vi else ""

        # 持久化到 YAML
        self._app.config_mgr.save_query_module(mod_id, new_cfg)

        # 内存即时生效
        if enabled:
            self._app._reinit_single_module(mod_id, new_cfg)
        else:
            self._app.query_modules.pop(mod_id, None)

        self._current_mod_id = mod_id
        self._is_new = False

        # 刷新列表并重新选中
        self.refresh()
        for i in range(self._list.count()):
            if self._list.item(i).data(Qt.UserRole) == mod_id:
                self._list.setCurrentRow(i)
                break

        QMessageBox.information(self, _t("保存成功"), _t("查询模块「{mod_id}」已保存并即时生效。").format(mod_id=mod_id))

    def _on_test(self) -> None:
        """用表单当前参数实例化后端并测试连接"""
        from core.query_base import BACKEND_REGISTRY
        b_type = self._type_combo.currentData() or ""
        cls = BACKEND_REGISTRY.get(b_type)
        if cls is None:
            self._test_label.setText(f"❌ 后端类型「{b_type}」未在插件注册表中，无法测试")
            self._test_label.setStyleSheet("color:#c62828; font-size:11px;")
            return

        # 从表单收集测试参数
        test_cfg: dict = {"type": b_type}
        if b_type in self._URL_TYPES:
            test_cfg["base_url"] = self._v_url.text().strip()
            test_cfg["timeout"] = self._v_timeout.value()
        elif b_type in ("pg_plan", "pg_spare_parts"):
            test_cfg["timeout"] = self._pg_timeout.value()
            test_cfg["database"] = {
                "host":     self._pg_host.text().strip(),
                "port":     self._pg_port.value(),
                "dbname":   self._pg_dbname.text().strip(),
                "user":     self._pg_user.text().strip(),
                "password": self._pg_pass.text(),
            }
        elif b_type == "excel_file":
            test_cfg["sources"] = [
                self._excel_src_list.item(i).data(Qt.UserRole)
                for i in range(self._excel_src_list.count())
                if self._excel_src_list.item(i).data(Qt.UserRole)
            ]
            test_cfg["has_header"] = self._excel_has_header.isChecked()
            checked_sheets = []
            for i in range(self._excel_sheet_list.count()):
                item = self._excel_sheet_list.item(i)
                if item.checkState() == Qt.Checked:
                    checked_sheets.append(item.text())
            sheet_val = ",".join(checked_sheets) if checked_sheets else "0"
            test_cfg["sheet"] = sheet_val
            test_cfg["max_rows"] = self._excel_max_rows.value()
            columns = []
            for r in range(self._excel_col_table.rowCount()):
                widget = self._excel_col_table.cellWidget(r, 0)
                is_checked = True
                if widget:
                    cb = widget.findChild(QCheckBox)
                    if cb:
                        is_checked = cb.isChecked()
                if not is_checked:
                    continue
                ci = self._excel_col_table.item(r, 1)
                fi = self._excel_col_table.item(r, 2)
                col_val   = ci.text().strip() if ci else ""
                field_val = fi.text().strip() if fi else ""
                if col_val:
                    columns.append({"col": col_val, "field": field_val or col_val})
            test_cfg["columns"] = columns

        elif b_type == "qdrant_vector":
            test_cfg.update({
                "qdrant_host":        self._qd_host.text().strip(),
                "qdrant_port":        self._qd_port.value(),
                "qdrant_collection":  self._qd_collection.text().strip(),
                "qdrant_api_key":     self._qd_api_key.text(),
                "qdrant_prefer_grpc": self._qd_prefer_grpc.isChecked(),
                "embed_base_url":     self._bge_url.text().strip(),
                "embed_api_key":      self._bge_api_key.text(),
                "embed_model":        self._bge_model.text().strip() or "BAAI/bge-m3",
                "embed_timeout":      self._bge_timeout.value(),
                "search_mode":        self._qd_mode.currentData() or "hybrid",
                "score_threshold":    self._qd_threshold.value(),
                "limit":              self._qd_limit.value(),
                "timeout":            self._qd_timeout.value(),
            })
        elif b_type == "sql_table":
            test_cfg.update({
                "db_type":  self._sql_db_type.currentData() or "postgresql",
                "host":     self._sql_pg_host.text().strip(),
                "port":     self._sql_pg_port.value(),
                "dbname":   self._sql_pg_dbname.text().strip(),
                "user":     self._sql_pg_user.text().strip(),
                "password": self._sql_pg_pass.text(),
                "db_path":  self._sql_db_path.text().strip(),
                "timeout":  self._sql_timeout.value(),
                "table":    self._sql_table_combo.currentText().strip(),
            })
        else:
            for r in range(self._kv_table.rowCount()):
                ki = self._kv_table.item(r, 0)
                vi = self._kv_table.item(r, 1)
                if ki and ki.text().strip():
                    test_cfg[ki.text().strip()] = vi.text().strip() if vi else ""

        self._test_btn.setEnabled(False)
        self._test_label.setText(_t("⏳ 正在测试连接…"))
        self._test_label.setStyleSheet("color:#555; font-size:11px;")

        self._test_thread = _ConnectTestThread(cls, test_cfg)
        self._test_thread.result.connect(self._on_test_result)
        self._test_thread.start()

    def _on_test_result(self, success: bool, msg: str) -> None:
        self._test_btn.setEnabled(True)
        if success:
            self._test_label.setText(f"✅ {msg}")
            self._test_label.setStyleSheet("color:#2e7d32; font-size:11px;")
            # 测试成功后：如果是 qdrant_vector 且字段列表为空，自动触发字段发现
            b_type = self._type_combo.currentData() or ""
            if b_type == "qdrant_vector" and self._qd_field_list.count() == 0:
                self._on_qd_fetch_fields()
            # 测试成功后：如果是 sql_table 且表列表为空，自动获取表列表
            elif b_type == "sql_table" and self._sql_table_combo.count() <= 1:
                self._sql_fetch_tables()
        else:
            self._test_label.setText(f"❌ {msg}")
            self._test_label.setStyleSheet("color:#c62828; font-size:11px;")

    # ── 键值对辅助 ───────────────────────────────────────────────

    def _kv_add_row(self, key: str = "", value: str = "") -> None:
        r = self._kv_table.rowCount()
        self._kv_table.insertRow(r)
        self._kv_table.setItem(r, 0, QTableWidgetItem(key))
        self._kv_table.setItem(r, 1, QTableWidgetItem(value))

    def _kv_del_row(self) -> None:
        r = self._kv_table.currentRow()
        if r >= 0:
            self._kv_table.removeRow(r)

    # ── Qdrant 辅助方法 ──────────────────────────────────────────

    def _on_qd_fetch_fields(self) -> None:
        """
        临时建立 Qdrant 连接，scroll 取少量点，自动发现 payload 字段，
        并填充到 _qd_field_list。
        """
        from PyQt5.QtWidgets import QApplication
        qd_host = self._qd_host.text().strip()
        qd_coll = self._qd_collection.text().strip()
        if not qd_host or not qd_coll:
            QMessageBox.warning(self, "提示", "请先填写 Qdrant 主机和数据集名称")
            return

        self._qd_fetch_fields_btn.setEnabled(False)
        self._qd_field_status.setText(_t("⏳ 正在连接 Qdrant 获取字段…"))
        QApplication.processEvents()

        try:
            from qdrant_client import QdrantClient
            cli = QdrantClient(
                host=qd_host,
                port=self._qd_port.value(),
                api_key=self._qd_api_key.text() or None,
                https=False,
                prefer_grpc=self._qd_prefer_grpc.isChecked(),
                timeout=self._qd_timeout.value(),
            )
            results, _ = cli.scroll(
                collection_name=qd_coll,
                limit=10,
                with_payload=True,
                with_vectors=False,
            )
            field_set: set[str] = set()
            for point in results:
                if point.payload:
                    field_set.update(point.payload.keys())

            # 记录当前已勾选字段，重建列表后尽量保持勾选
            old_checked = {
                self._qd_field_list.item(i).text()
                for i in range(self._qd_field_list.count())
                if self._qd_field_list.item(i).checkState() == Qt.Checked
            }

            self._qd_field_list.clear()
            # score 始终排第一
            score_item = QListWidgetItem("score")
            score_item.setFlags(score_item.flags() | Qt.ItemIsUserCheckable)
            score_item.setCheckState(Qt.Checked if "score" in old_checked else Qt.Unchecked)
            self._qd_field_list.addItem(score_item)
            for f in sorted(field_set):
                item = QListWidgetItem(f)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Checked if f in old_checked else Qt.Unchecked)
                self._qd_field_list.addItem(item)
            self._adjust_qd_field_list_height()

            n = len(field_set) + 1   # +1 for score
            self._qd_field_status.setText(
                _t("共 {n} 个字段（含 score）。✅ 勾选后仅返回该字段，无勾选则返回所有。").format(n=n)
            )
        except ImportError:
            self._qd_field_status.setText("❌ 未安装 qdrant-client，请运行: pip install qdrant-client")
        except Exception as exc:
            self._qd_field_status.setText(f"❌ 获取失败：{exc}")
        finally:
            self._qd_fetch_fields_btn.setEnabled(True)

    def _qd_field_check_all(self) -> None:
        """勾选列表中所有字段"""
        for i in range(self._qd_field_list.count()):
            self._qd_field_list.item(i).setCheckState(Qt.Checked)

    def _qd_field_uncheck_all(self) -> None:
        """取消勾选列表中所有字段"""
        for i in range(self._qd_field_list.count()):
            self._qd_field_list.item(i).setCheckState(Qt.Unchecked)

    def _qd_field_remove_selected(self) -> None:
        """移除当前高亮（点击选中）的字段行"""
        row = self._qd_field_list.currentRow()
        if row >= 0:
            self._qd_field_list.takeItem(row)
            self._adjust_qd_field_list_height()
            n = self._qd_field_list.count()
            checked = sum(
                1 for i in range(n)
                if self._qd_field_list.item(i).checkState() == Qt.Checked
            )
            self._qd_field_status.setText(
                _t("共 {n} 个字段，已勾选 {checked} 个。").format(n=n, checked=checked) if n else _t("（字段列表已清空，将返回全部字段）")
            )

    def _qd_filter_add_row(self, field: str = "", value: str = "") -> None:
        r = self._qd_filter_table.rowCount()
        self._qd_filter_table.insertRow(r)
        self._qd_filter_table.setItem(r, 0, QTableWidgetItem(field))
        self._qd_filter_table.setItem(r, 1, QTableWidgetItem(value))

    def _qd_filter_del_row(self) -> None:
        r = self._qd_filter_table.currentRow()
        if r >= 0:
            self._qd_filter_table.removeRow(r)

    # ── Excel 数据源辅助方法 ─────────────────────────────────

    def _excel_add_files(self) -> None:
        """弹出文件对话框，支持多选 Excel 文件"""
        from PyQt5.QtWidgets import QFileDialog
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx *.xlsm *.xltx *.xltm);;所有文件 (*)"
        )
        for p in paths:
            src  = {"type": "file", "path": p}
            item = QListWidgetItem(f"📄 {p}")
            item.setData(Qt.UserRole, src)
            self._excel_src_list.addItem(item)

    def _excel_add_directory(self) -> None:
        """弹出目录对话框，并询问是否递归子目录"""
        from PyQt5.QtWidgets import QFileDialog, QInputDialog
        d = QFileDialog.getExistingDirectory(self, _t("选择 Excel 文件所在目录"))
        if not d:
            return
        choice, ok = QInputDialog.getItem(
            self, _t("遍历选项"), _t("是否递归遍历子目录？"),
            [_t("否（仅当前目录）"), _t("是（包含所有子目录）")], 0, False
        )
        if not ok:
            return
        recursive = choice.startswith(_t("是"))
        src   = {"type": "directory", "path": d, "recursive": recursive}
        label = f"📁 {d}{'  （含子目录）' if recursive else ''}"
        item  = QListWidgetItem(label)
        item.setData(Qt.UserRole, src)
        self._excel_src_list.addItem(item)

    def _excel_del_source(self) -> None:
        """删除选中的数据源"""
        for item in self._excel_src_list.selectedItems():
            self._excel_src_list.takeItem(self._excel_src_list.row(item))

    def _excel_add_col_row(self, col: str = "", field: str = "", checked: bool = True) -> None:
        """在列映射表末尾插入一行"""
        r = self._excel_col_table.rowCount()
        self._excel_col_table.insertRow(r)
        
        # 第 0 列居中复选框
        cb = QCheckBox()
        cb.setChecked(checked)
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.addWidget(cb)
        layout.setAlignment(Qt.AlignCenter)
        layout.setContentsMargins(0, 0, 0, 0)
        self._excel_col_table.setCellWidget(r, 0, widget)

        self._excel_col_table.setItem(r, 1, QTableWidgetItem(col))
        self._excel_col_table.setItem(r, 2, QTableWidgetItem(field))
        self._adjust_excel_col_table_height()

    def _excel_del_col_row(self) -> None:
        """删除列映射表中选中的行"""
        r = self._excel_col_table.currentRow()
        if r >= 0:
            self._excel_col_table.removeRow(r)
            self._adjust_excel_col_table_height()

    def _adjust_excel_col_table_height(self) -> None:
        """动态调节列映射表的高度，使其完全适应内容行数，避免受外层滚动区域 size hint 限制而缩回"""
        rows = self._excel_col_table.rowCount()
        # 表头高度
        header_height = self._excel_col_table.horizontalHeader().height() or 30
        # 所有行的总高
        row_heights = sum(self._excel_col_table.rowHeight(r) or 28 for r in range(rows))
        
        # 计算出所需的基本高度
        total_height = header_height + row_heights + self._excel_col_table.frameWidth() * 2 + 10
        
        # 我们动态设定其最小高度，使其强制扩展，解决 QScrollArea 的 size hint 导致的限高 bug
        # 不低于 220 像素以呈现良好的视觉基础，上限设定在 800 像素以防过度膨胀
        target_height = max(220, min(total_height, 800))
        self._excel_col_table.setMinimumHeight(target_height)
        self._excel_col_table.updateGeometry()

    def _adjust_sql_field_table_height(self) -> None:
        """动态调节 SQL 字段配置表的高度，避免受外层滚动区域 size hint 限制而缩回"""
        rows = self._sql_field_table.rowCount()
        header_height = self._sql_field_table.horizontalHeader().height() or 30
        row_heights = sum(self._sql_field_table.rowHeight(r) or 28 for r in range(rows))
        total_height = header_height + row_heights + self._sql_field_table.frameWidth() * 2 + 10
        target_height = max(220, min(total_height, 800))
        self._sql_field_table.setMinimumHeight(target_height)
        self._sql_field_table.updateGeometry()

    def _adjust_qd_field_list_height(self) -> None:
        """动态调节 Qdrant 返回字段列表的高度，避免受外层滚动区域 size hint 限制而缩回"""
        count = self._qd_field_list.count()
        # 列表每项高度大约 24 像素
        item_heights = count * 24
        total_height = item_heights + self._qd_field_list.frameWidth() * 2 + 10
        target_height = max(200, min(total_height, 600))
        self._qd_field_list.setMinimumHeight(target_height)
        self._qd_field_list.updateGeometry()

    def _excel_autodetect_cols(self) -> None:
        """
        根据当前选中的数据源文件自动填充列映射表。
        """
        if self._excel_src_list.count() == 0:
            QMessageBox.information(self, "提示", "请先添加至少一个数据源")
            return

        selected = self._excel_src_list.selectedItems()
        target_item = selected[0] if selected else (self._excel_src_list.item(0) if self._excel_src_list.count() > 0 else None)
        filepath = self._get_filepath_from_item(target_item)
        if filepath is None:
            QMessageBox.warning(self, _t("错误"), _t("无法找到可读取的 Excel 文件"))
            return

        from pathlib import Path
        first_file = Path(filepath)

        # 获取首个勾选的工作表名称作为检测对象
        sheet_ident = None
        for i in range(self._excel_sheet_list.count()):
            item = self._excel_sheet_list.item(i)
            if item.checkState() == Qt.Checked:
                sheet_ident = item.text()
                break
        if not sheet_ident:
            if self._excel_sheet_list.count() > 0:
                sheet_ident = self._excel_sheet_list.item(0).text()
            else:
                sheet_ident = "0"

        # 兼容处理
        if isinstance(sheet_ident, str) and sheet_ident.isdigit():
            sv_val = int(sheet_ident)
        else:
            sv_val = sheet_ident or "0"

        first_row  = None

        # ── 尝试 calamine（Rust，优先） ──────────────────────────
        try:
            from python_calamine import CalamineWorkbook
            wb          = CalamineWorkbook.from_path(str(first_file))
            sheet_names = wb.sheet_names

            if isinstance(sv_val, int):
                ws = wb.get_sheet_by_index(min(sv_val, len(sheet_names) - 1))
            elif isinstance(sv_val, str) and sv_val in sheet_names:
                ws = wb.get_sheet_by_name(sv_val)
            else:
                try:
                    idx = int(sv_val)
                    ws = wb.get_sheet_by_index(min(idx, len(sheet_names) - 1))
                except ValueError:
                    ws = wb.get_sheet_by_index(0)

            rows = ws.to_python(skip_empty_area=False)
            if rows:
                first_row = rows[0]

        except ImportError:
            # calamine 不可用，降级到 openpyxl
            try:
                import openpyxl
                wb = openpyxl.load_workbook(
                    str(first_file), read_only=True, data_only=True
                )
                if isinstance(sv_val, int):
                    ws = wb.worksheets[min(sv_val, len(wb.worksheets) - 1)]
                elif isinstance(sv_val, str) and sv_val in wb.sheetnames:
                    ws = wb[sv_val]
                else:
                    try:
                        idx = int(sv_val)
                        ws = wb.worksheets[min(idx, len(wb.worksheets) - 1)]
                    except ValueError:
                        ws = wb.active
                for row in ws.iter_rows(max_row=1, values_only=True):
                    first_row = row
                    break
                wb.close()
            except ImportError:
                QMessageBox.critical(
                    self, "错误",
                    "缺少 Excel 读取库。\n请运行: pip install python-calamine"
                )
                return
            except Exception as exc:
                QMessageBox.critical(self, "检测失败（openpyxl）", str(exc))
                return

        except Exception as exc:
            QMessageBox.critical(self, "检测失败（calamine）", str(exc))
            return

        if not first_row:
            QMessageBox.information(self, "提示", "工作表第一行为空")
            return

        has_header = self._excel_has_header.isChecked()
        self._excel_col_table.setRowCount(0)
        detected_fields = []
        for i, val in enumerate(first_row):
            col_letter = _ui_index_to_col_letter(i)
            if has_header:
                field_name = str(val).strip() if (val is not None and str(val).strip()) else col_letter
            else:
                field_name = col_letter
            self._excel_add_col_row(col_letter, field_name, checked=True)
            detected_fields.append(f"  - {col_letter}: {field_name}")

        detected_summary = "\n".join(detected_fields)
        QMessageBox.information(
            self, _t("自动检测完成"),
            _t("从文件「{filename}」（工作表: {sheet_ident}）检测到 {col_len} 列，已全部自动添加并勾选：\n\n{detected_summary}").format(
                filename=first_file.name, sheet_ident=sheet_ident, col_len=len(first_row), detected_summary=detected_summary)
        )

    def _get_filepath_from_item(self, item: QListWidgetItem) -> str | None:
        """从 QListWidgetItem 提取 Excel 数据源的有效文件路径"""
        if not item:
            return None
        src = item.data(Qt.UserRole) or {}
        src_type = src.get("type", "file")
        raw_path = src.get("path", "")
        if not raw_path:
            return None
        
        from pathlib import Path
        _ALL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".ods"}
        p = Path(raw_path)
        if src_type == "file" and p.is_file():
            return str(p)
        elif src_type == "directory" and p.is_dir():
            for fp in sorted(p.iterdir()):
                if fp.is_file() and fp.suffix.lower() in _ALL_SUFFIXES:
                    return str(fp)
        return None

    def _on_excel_src_selection_changed(self) -> None:
        """当数据源选择项改变时，自动重新加载对应文件的工作表列表"""
        selected = self._excel_src_list.selectedItems()
        if selected:
            # 自动加载并切换工作表，不弹成功提示框
            self._excel_load_sheets(target_item=selected[0], show_success_msg=False)

    def _excel_load_sheets(self, target_item: QListWidgetItem = None, show_success_msg: bool = True) -> None:
        """读取指定或当前选中数据源文件的所有工作表并填充到列表中"""
        if self._excel_src_list.count() == 0:
            if show_success_msg:
                QMessageBox.information(self, "提示", "请先添加至少一个数据源")
            return

        # 优先读取传入的 target_item，否则寻找当前选中的，若没有则用首个
        if not target_item:
            selected = self._excel_src_list.selectedItems()
            target_item = selected[0] if selected else self._excel_src_list.item(0)

        filepath = self._get_filepath_from_item(target_item)
        if filepath is None:
            if show_success_msg:
                QMessageBox.warning(self, _t("错误"), _t("无法找到可读取的 Excel 文件"))
            return

        loaded_sheets = []
        try:
            from python_calamine import CalamineWorkbook
            wb = CalamineWorkbook.from_path(filepath)
            loaded_sheets = wb.sheet_names
        except ImportError:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True)
                loaded_sheets = wb.sheetnames
                wb.close()
            except Exception as e:
                if show_success_msg:
                    QMessageBox.critical(self, "错误", f"读取工作表失败: {e}")
                return
        except Exception as e:
            if show_success_msg:
                QMessageBox.critical(self, "错误", f"读取工作表失败: {e}")
            return

        self._excel_sheet_list.clear()
        for idx, s_name in enumerate(loaded_sheets):
            item = QListWidgetItem(s_name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            # 默认勾选第一个工作表
            item.setCheckState(Qt.Checked if idx == 0 else Qt.Unchecked)
            self._excel_sheet_list.addItem(item)
            
        if show_success_msg:
            from pathlib import Path
            QMessageBox.information(self, _t("成功"), _t("从「{filename}」成功加载了 {count} 个工作表。").format(filename=Path(filepath).name, count=len(loaded_sheets)))

    def _on_excel_sheet_item_clicked(self, item: QListWidgetItem) -> None:
        """当用户双击工作表项时，极速读取其前 20 行并显示预览窗口"""
        sheet_name = item.text()
        
        # 获取选中的或首个源文件
        if self._excel_src_list.count() == 0:
            return

        selected = self._excel_src_list.selectedItems()
        target_item = selected[0] if selected else self._excel_src_list.item(0)

        filepath = self._get_filepath_from_item(target_item)
        if filepath is None:
            return

        preview_rows = []
        try:
            from python_calamine import CalamineWorkbook
            wb = CalamineWorkbook.from_path(filepath)
            if sheet_name in wb.sheet_names:
                ws = wb.get_sheet_by_name(sheet_name)
                # 读取前 20 行
                raw = ws.to_python(skip_empty_area=False)
                preview_rows = raw[:20]
        except ImportError:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for idx, row in enumerate(ws.iter_rows(values_only=True)):
                        preview_rows.append(row)
                        if idx >= 19:
                            break
                wb.close()
            except Exception as e:
                logger.error("预览工作表读取失败: %s", e)
        except Exception as e:
            logger.error("预览工作表读取失败: %s", e)

        if not preview_rows:
            QMessageBox.warning(self, _t("预览失败"), _t("未能从工作表中读取到任何数据。"))
            return

        # 转换为字符串列表以便在表格中展示
        formatted_rows = []
        for row in preview_rows:
            formatted_row = [str(cell) if cell is not None else "" for cell in row]
            formatted_rows.append(formatted_row)

        self._show_preview_dialog(sheet_name, formatted_rows)

    def _show_preview_dialog(self, sheet_name: str, rows: list[list[str]]) -> None:
        """弹出一个高级的独立表格对话框展示工作表的前 20 行网格内容，表头使用 ASCII 列标"""
        preview_dialog = QDialog(self)
        preview_dialog.setWindowTitle(_t(f"工作表预览 — {sheet_name} (前20行)"))
        preview_dialog.resize(850, 500)
        preview_dialog.setWindowFlags(preview_dialog.windowFlags() | Qt.WindowMaximizeButtonHint)

        dlg_layout = QVBoxLayout(preview_dialog)
        dlg_layout.setContentsMargins(12, 12, 12, 12)
        dlg_layout.setSpacing(8)

        # 头部说明
        header_label = QLabel(_t(f"正在浏览工作表「{sheet_name}」的数据样例："))
        header_label.setStyleSheet("font-weight: bold; font-size: 13px; color: #333;")
        dlg_layout.addWidget(header_label)

        # QTableWidget 展示数据
        max_cols = max(len(row) for row in rows) if rows else 0
        table = QTableWidget(len(rows), max_cols)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        
        # 设置列标题为 Excel 字母列标 A, B, C...
        col_headers = [_ui_index_to_col_letter(i) for i in range(max_cols)]
        table.setHorizontalHeaderLabels(col_headers)
        
        # 填充数据
        for r_idx, row in enumerate(rows):
            for c_idx, cell_val in enumerate(row):
                table.setItem(r_idx, c_idx, QTableWidgetItem(cell_val))
                
        # 自动调整列宽以适应内容，但限制最大宽度防止过宽
        table.resizeColumnsToContents()
        for c in range(max_cols):
            if table.columnWidth(c) > 200:
                table.setColumnWidth(c, 200)

        dlg_layout.addWidget(table)

        # 底部操作栏：提供一键关闭按钮
        bottom_layout = QHBoxLayout()
        bottom_layout.addStretch()
        close_btn = QPushButton(_t("关闭预览"))
        close_btn.clicked.connect(preview_dialog.accept)
        close_btn.setStyleSheet("padding: 6px 20px; font-weight: bold;")
        bottom_layout.addWidget(close_btn)
        
        dlg_layout.addLayout(bottom_layout)
        preview_dialog.exec_()

    # ── SQL 辅助方法 ──────────────────────────────────────────────

    def _sql_on_dbtype_changed(self) -> None:
        """根据选择的数据库类型，显示/隐藏对应的连接参数控件"""
        db_type = self._sql_db_type.currentData() or "postgresql"
        is_pg   = (db_type == "postgresql")
        is_sl   = (db_type == "sqlite")

        # 找到 FormLayout 行标签并控制可见性
        # 通过直接操作各控件及其父 row 的可见性
        for w in self._sql_pg_widgets:
            # 找到 w 在 QFormLayout 中的行，只能通过 parentWidget() 路径
            w.setVisible(is_pg)
            # 同步标签
            parent_fl = w.parent()
            if parent_fl is not None:
                lbl = parent_fl.findChild(type(QLabel()), "", Qt.FindDirectChildrenOnly)

        # 简单方案：直接设置可见
        self._sql_pg_host.setVisible(is_pg)
        self._sql_pg_port.setVisible(is_pg)
        self._sql_pg_dbname.setVisible(is_pg)
        self._sql_pg_user.setVisible(is_pg)
        self._sql_pg_pass.setVisible(is_pg)
        self._sql_db_path.setVisible(is_sl)

    def _sql_browse_file(self) -> None:
        """浏览选择 SQLite 数据库文件"""
        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 SQLite 数据库文件", "",
            "SQLite 数据库 (*.db *.sqlite *.sqlite3 *.s3db);;所有文件 (*)"
        )
        if path:
            self._sql_db_path.setText(path)

    def _sql_build_test_cfg(self) -> dict:
        """从表单收集用于临时连接的配置 dict"""
        return {
            "db_type":  self._sql_db_type.currentData() or "postgresql",
            "host":     self._sql_pg_host.text().strip(),
            "port":     self._sql_pg_port.value(),
            "dbname":   self._sql_pg_dbname.text().strip(),
            "user":     self._sql_pg_user.text().strip(),
            "password": self._sql_pg_pass.text(),
            "db_path":  self._sql_db_path.text().strip(),
            "timeout":  self._sql_timeout.value(),
        }

    def _sql_fetch_tables(self) -> None:
        """临时建立连接，获取所有表/视图名，填充到表下拉"""
        from PyQt5.QtWidgets import QApplication
        self._sql_fetch_tables_btn.setEnabled(False)
        self._sql_table_combo.setPlaceholderText(_t("⏳ 正在获取表列表…"))
        QApplication.processEvents()

        try:
            from core.query_backends.sql_table_query import get_tables_from_cfg
            cfg    = self._sql_build_test_cfg()
            tables = get_tables_from_cfg(cfg)

            cur_text = self._sql_table_combo.currentText().strip()
            self._sql_table_combo.blockSignals(True)
            self._sql_table_combo.clear()
            for t in tables:
                self._sql_table_combo.addItem(t)
            if cur_text:
                idx = self._sql_table_combo.findText(cur_text)
                if idx >= 0:
                    self._sql_table_combo.setCurrentIndex(idx)
                else:
                    self._sql_table_combo.setCurrentText(cur_text)
            self._sql_table_combo.blockSignals(False)
            self._sql_table_combo.setPlaceholderText("")
            n = len(tables)
            self._test_label.setText(_t("✅ 获取到 {n} 个表/视图").format(n=n))
            self._test_label.setStyleSheet("color:#2e7d32; font-size:11px;")
        except Exception as exc:
            self._sql_table_combo.setPlaceholderText(_t("❌ 获取失败，请检查连接信息"))
            self._test_label.setText(_t("❌ 获取表列表失败：{err}").format(err=str(exc)[:80]))
            self._test_label.setStyleSheet("color:#c62828; font-size:11px;")
        finally:
            self._sql_fetch_tables_btn.setEnabled(True)

    def _sql_on_table_changed(self, text: str) -> None:
        """表下拉变更时，自动获取字段（如果字段列表为空）"""
        if text.strip() and self._sql_field_table.rowCount() == 0:
            self._sql_fetch_fields()

    def _sql_fetch_fields(self) -> None:
        """获取当前选定表的所有字段，填充到字段配置表"""
        from PyQt5.QtWidgets import QApplication
        table = self._sql_table_combo.currentText().strip()
        if not table:
            QMessageBox.warning(self, _t("提示"), _t("请先选择目标表/视图"))
            return

        self._sql_fetch_fields_btn.setEnabled(False)
        QApplication.processEvents()
        try:
            from core.query_backends.sql_table_query import get_fields_from_cfg
            cfg    = self._sql_build_test_cfg()
            fields = get_fields_from_cfg(cfg, table)

            # 记录当前已有的勾选状态，避免重建时丢失
            old_ret: set[str]  = set()
            old_srch: set[str] = set()
            for r in range(self._sql_field_table.rowCount()):
                fn_item = self._sql_field_table.item(r, 0)
                cb_ret  = self._sql_field_table.cellWidget(r, 1)
                cb_srch = self._sql_field_table.cellWidget(r, 2)
                fn = fn_item.text().strip() if fn_item else ""
                if fn:
                    if cb_ret  and cb_ret.isChecked():
                        old_ret.add(fn)
                    if cb_srch and cb_srch.isChecked():
                        old_srch.add(fn)

            self._sql_field_table.setRowCount(0)
            for f in fields:
                ret_default  = (f in old_ret)  if old_ret  else True
                srch_default = (f in old_srch) if old_srch else False
                self._sql_add_field_row(f, ret_default, srch_default)

            self._test_label.setText(f"✅ 获取到 {len(fields)} 个字段")
            self._test_label.setStyleSheet("color:#2e7d32; font-size:11px;")
        except Exception as exc:
            self._test_label.setText(f"❌ 获取字段失败：{str(exc)[:80]}")
            self._test_label.setStyleSheet("color:#c62828; font-size:11px;")
        finally:
            self._sql_fetch_fields_btn.setEnabled(True)

    def _sql_add_field_row(
        self, field: str, ret_checked: bool = True, srch_checked: bool = False
    ) -> None:
        """在字段配置表末尾插入一行"""
        r = self._sql_field_table.rowCount()
        self._sql_field_table.insertRow(r)

        fn_item = QTableWidgetItem(field)
        fn_item.setFlags(fn_item.flags() & ~Qt.ItemIsEditable)
        self._sql_field_table.setItem(r, 0, fn_item)

        cb_ret = QCheckBox()
        cb_ret.setChecked(ret_checked)
        cb_ret.setStyleSheet("margin-left:12px;")
        self._sql_field_table.setCellWidget(r, 1, cb_ret)

        cb_srch = QCheckBox()
        cb_srch.setChecked(srch_checked)
        cb_srch.setStyleSheet("margin-left:12px;")
        self._sql_field_table.setCellWidget(r, 2, cb_srch)
        self._adjust_sql_field_table_height()

    def _sql_field_check_col(self, col: int, checked: bool) -> None:
        """批量勾选/取消某列（1=返回，2=查询）的所有复选框"""
        for r in range(self._sql_field_table.rowCount()):
            cb = self._sql_field_table.cellWidget(r, col)
            if cb:
                cb.setChecked(checked)


# ══════════════════════════════════════════════════════════════════
# 字段选择器对话框（辅助工具）
# ══════════════════════════════════════════════════════════════════

class _FieldPickerDialog(QDialog):
    """从字段列表中多选字段"""
    def __init__(self, fields: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle(_t("选择字段"))
        self.setMinimumSize(260, 380)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(_t("双击或选中后点确定：")))
        self._list = QListWidget()
        self._list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        for f in fields:
            self._list.addItem(f)
        self._list.itemDoubleClicked.connect(self.accept)
        layout.addWidget(self._list)
        bar = QHBoxLayout()
        bar.addStretch()
        ok = QPushButton(_t("确定"))
        ok.clicked.connect(self.accept)
        ca = QPushButton(_t("取消"))
        ca.clicked.connect(self.reject)
        bar.addWidget(ok); bar.addWidget(ca)
        layout.addLayout(bar)

    def selected_fields(self) -> list[str]:
        return [it.text() for it in self._list.selectedItems()]


# ══════════════════════════════════════════════════════════════════
# 映射组编辑对话框（选项卡 + 排序 + 字段选择）
# ══════════════════════════════════════════════════════════════════

class MappingGroupDialog(QDialog):
    """创建或编辑映射组的弹窗"""

    def __init__(self, app_controller: Application, group_id: str = "", group=None, parent=None):
        super().__init__(parent)
        self._app = app_controller
        self._group_id = group_id  # 编辑时保留原 ID；新建时为空（保存时自动生成）
        self.setWindowTitle(_t("编辑映射组") if group_id else _t("新建映射组"))
        self.setMinimumSize(780, 640)
        self._setup_ui(group)

    # ── 可用字段 ─────────────────────────────────────────────────

    def _get_module_fields(self) -> list[str]:
        mod_id = self._module_combo.currentData()
        # ① 优先从运行中且已就绪的模块实例读取（已连接 Qdrant 时有完整字段）
        qm = self._app.query_modules.get(mod_id)
        if qm and hasattr(qm, "get_available_fields"):
            # 只有 status==ready 时 live 字段才可信
            status_ok = not hasattr(qm, "get_status") or qm.get_status() == "ready"
            if status_ok:
                live_fields = [f[0] for f in qm.get_available_fields()]
                if live_fields:
                    return sorted(live_fields)
        # ② 兜底：从 config 中读取 return_fields（模块未启动/未连接时）
        cfg = self._app.config_mgr.get_query_modules().get(mod_id, {})
        cfg_fields = cfg.get("return_fields") or []
        if cfg_fields:
            return sorted(cfg_fields)
        return []

    def _refresh_field_list(self) -> None:
        """刷新左侧字段列表（显示列 Tab）"""
        self._field_list.clear()
        for f in self._get_module_fields():
            self._field_list.addItem(f)

    def _refresh_output_combos(self) -> None:
        """刷新输出映射表中的字段 ComboBox"""
        fields = self._get_module_fields()
        for r in range(self._out_table.rowCount()):
            w = self._out_table.cellWidget(r, 0)
            if isinstance(w, QComboBox):
                cur = w.currentText()
                w.blockSignals(True)
                w.clear()
                for f in fields:
                    w.addItem(f)
                idx = w.findText(cur)
                w.setCurrentIndex(idx if idx >= 0 else 0)
                w.blockSignals(False)

    def _refresh_input_query_combos(self) -> None:
        """刷新输入映射表中“查询字段”的 ComboBox"""
        fields = self._get_module_fields()
        for r in range(self._in_table.rowCount()):
            w = self._in_table.cellWidget(r, 0)
            if isinstance(w, QComboBox):
                cur_val = w.currentData() or w.currentText()
                w.blockSignals(True)
                w.clear()
                w.addItem("（无）", "")
                for f in fields:
                    w.addItem(f, f)
                idx = w.findData(cur_val)
                if idx >= 0:
                    w.setCurrentIndex(idx)
                elif cur_val:
                    w.addItem(cur_val, cur_val)
                    w.setCurrentIndex(w.count() - 1)
                w.blockSignals(False)

    def _on_module_changed(self) -> None:
        self._refresh_field_list()
        self._refresh_output_combos()
        self._refresh_input_query_combos()


    # ── UI 构建 ──────────────────────────────────────────────────

    def _setup_ui(self, group) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 8)
        layout.setSpacing(8)

        # ── 基本信息 ─────────────────────────────────────────────
        basic = QGroupBox(_t("基本信息"))
        bf = QFormLayout(basic)
        bf.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self._name_input = QLineEdit(group.name if group else "")
        self._name_input.setPlaceholderText(_t("显示名称"))
        bf.addRow(_t("名称:"), self._name_input)

        # 查询模块 + 最大结果数
        mod_row = QHBoxLayout()
        self._module_combo = QComboBox()
        for mid, qm_cfg in self._app.config_mgr.get_query_modules().items():
            name = qm_cfg.get("name", mid)
            self._module_combo.addItem(f"{name} [{mid}]", mid)
        if group:
            idx = self._module_combo.findData(group.query_module)
            if idx >= 0:
                self._module_combo.setCurrentIndex(idx)
        self._module_combo.currentIndexChanged.connect(self._on_module_changed)
        mod_row.addWidget(self._module_combo, 1)
        mod_row.addSpacing(16)
        mod_row.addWidget(QLabel(_t("最大结果数:")))
        self._limit_spin = QSpinBox()
        self._limit_spin.setRange(1, 2000)
        self._limit_spin.setValue(group.query_params.get("limit", 20) if group else 20)
        mod_row.addWidget(self._limit_spin)
        bf.addRow(_t("查询模块:"), mod_row)

        # 交互模块下拉
        self._im_combo = QComboBox()
        self._im_combo.addItem(_t("（未设置）"), "")
        for im_id, im in self._app.config_mgr.get_interaction_modules().items():
            self._im_combo.addItem(f"{im.name} [{im_id}]", im_id)
        if group and group.interaction_module:
            idx = self._im_combo.findData(group.interaction_module)
            if idx >= 0:
                self._im_combo.setCurrentIndex(idx)
        self._im_combo.currentIndexChanged.connect(self._on_im_changed)
        bf.addRow(_t("交互模块:"), self._im_combo)

        layout.addWidget(basic)

        # ── 选项卡 ───────────────────────────────────────────────
        self._tabs = QTabWidget()
        self._in_table   = self._build_in_tab()
        self._disp_table, self._field_list = self._build_disp_tab()
        self._out_table  = self._build_out_tab()
        layout.addWidget(self._tabs, 1)

        # ── 底部按钮 ─────────────────────────────────────────────
        bar = QHBoxLayout()
        bar.addStretch()
        ok = QPushButton(_t("✔ 确定"))
        ok.setFixedWidth(90)
        ok.clicked.connect(self._on_ok)
        ca = QPushButton(_t("✘ 取消"))
        ca.setFixedWidth(90)
        ca.clicked.connect(self.reject)
        bar.addWidget(ok); bar.addWidget(ca)
        layout.addLayout(bar)

        # 填充已有数据
        if group:
            self._load_group(group)
        self._refresh_field_list()

    # ── Tab 工厂 ─────────────────────────────────────────────────

    def _build_in_tab(self) -> QTableWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.addWidget(QLabel(
            _t("定义每个查询字段的数据来源（三种模式互斜）：\n"
            "「 Excel列 + UI组件 」: 读 Excel → 自动填入对应组件 → 查询\n"
            "「 仅 Excel列 」: 读 Excel → 直接查询，不填入 UI\n"
            "「 仅 UI组件 」: Excel列留空，用户在界面手动输入后点查询")
        ))
        tbl = QTableWidget(0, 5)
        tbl.setHorizontalHeaderLabels([_t("查询字段"), _t("组件ID"), _t("Excel 列（逗号分隔）"), _t("拼接符"), _t("查询默认值")])
        h = tbl.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Stretch)
        h.setSectionResizeMode(1, QHeaderView.Stretch)
        h.setSectionResizeMode(2, QHeaderView.Stretch)
        h.setSectionResizeMode(3, QHeaderView.Fixed)
        h.setSectionResizeMode(4, QHeaderView.Stretch)
        tbl.setColumnWidth(3, 60)
        v.addWidget(tbl)
        v.addLayout(self._std_bar(tbl, [
            (_t("➕ 添加行"), lambda: self._in_mapping_row(tbl, "", "", "", " ", ""))
        ]))
        v.addStretch()
        self._tabs.addTab(w, _t("📥 输入映射"))
        return tbl

    def _build_disp_tab(self):
        w = QWidget()
        h_lay = QHBoxLayout(w)
        h_lay.setSpacing(6)

        # 左：字段列表
        lv = QVBoxLayout()
        lv.addWidget(QLabel(_t("可用字段\n(双击/批量添加):")))
        flist = QListWidget()
        flist.setMaximumWidth(170)
        flist.setSelectionMode(QAbstractItemView.ExtendedSelection)
        lv.addWidget(flist)
        add_all_btn = QPushButton(_t("→ 批量添加"))
        add_all_btn.clicked.connect(lambda: self._add_fields_to_disp(
            [flist.item(i).text() for i in range(flist.count())
             if flist.item(i).isSelected()] or
            [flist.item(i).text() for i in range(flist.count())]
        ))
        lv.addWidget(add_all_btn)
        h_lay.addLayout(lv)

        # 右：配置表
        rv = QVBoxLayout()
        rv.addWidget(QLabel(_t("显示列配置（↑↓调整显示顺序）:")))
        tbl = QTableWidget(0, 3)
        tbl.setHorizontalHeaderLabels([_t("结果字段"), _t("表头名称"), _t("列宽(px)")])
        hdr = tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        tbl.setColumnWidth(2, 70)
        rv.addWidget(tbl)
        rv.addLayout(self._std_bar(tbl, []))
        h_lay.addLayout(rv, 1)

        flist.itemDoubleClicked.connect(
            lambda item: self._add_fields_to_disp([item.text()])
        )

        self._tabs.addTab(w, _t("📊 显示列"))
        return tbl, flist

    def _build_out_tab(self) -> QTableWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.addWidget(QLabel(_t("定义双击结果行后将哪些字段写回 Excel（可留空不写入）：")))
        tbl = QTableWidget(0, 2)
        tbl.setHorizontalHeaderLabels([_t("结果字段"), _t("写入 Excel 列")])
        h = tbl.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Stretch)
        h.setSectionResizeMode(1, QHeaderView.Fixed)
        tbl.setColumnWidth(1, 100)
        v.addWidget(tbl)
        v.addLayout(self._std_bar(tbl, [
            (_t("➕ 从字段添加"), lambda: self._pick_output_fields(tbl))
        ]))
        v.addStretch()
        self._tabs.addTab(w, _t("📤 输出映射"))
        return tbl

    # ── 标准按钮栏（删除 + ↑↓） ─────────────────────────────────

    def _std_bar(self, tbl: QTableWidget, extras: list) -> QHBoxLayout:
        bar = QHBoxLayout()
        for lbl, slot in extras:
            b = QPushButton(lbl); b.clicked.connect(slot); bar.addWidget(b)
        d = QPushButton(_t("🗑 删除行"))
        d.clicked.connect(lambda: self._del_row(tbl))
        bar.addWidget(d)
        bar.addSpacing(8)
        u = QPushButton("↑"); u.setFixedWidth(30)
        u.clicked.connect(lambda: self._move_row(tbl, -1))
        dn = QPushButton("↓"); dn.setFixedWidth(30)
        dn.clicked.connect(lambda: self._move_row(tbl, +1))
        bar.addWidget(u); bar.addWidget(dn)
        bar.addStretch()
        return bar

    # ── 行操作工具 ───────────────────────────────────────────────

    def _plain_row(self, tbl: QTableWidget, vals: list) -> None:
        r = tbl.rowCount(); tbl.insertRow(r)
        for c, v in enumerate(vals):
            tbl.setItem(r, c, QTableWidgetItem(v))

    def _get_im_component_ids(self) -> list[str]:
        """获取当前选中的交互模块中定义的所有 UI 组件 ID"""
        im_id = self._im_combo.currentData()
        if not im_id:
            return []
        im = self._app.config_mgr.get_interaction_modules().get(im_id)
        if not im:
            return []

        ids = []
        if im.base_type == "display_panel":
            ids = [comp.field for comp in im.components if comp.field]
        elif im.base_type == "search_panel":
            ids = [sb.get("field") for sb in im.search_boxes if sb.get("field")]
        elif im.base_type == "popup_table":
            ids = [ctrl.query_field for ctrl in im.ui_controls if ctrl.query_field]
        return ids

    def _in_mapping_row(self, tbl: QTableWidget, qfield: str = "", ui_comp: str = "", cols: str = "", sep: str = " ", def_val: str = "") -> None:
        """在输入映射表中添加一行，其中第一列为查询字段下拉框（可编辑），第二列为组件ID下拉框，第五列为查询默认值"""
        r = tbl.rowCount()
        tbl.insertRow(r)

        # 列 0: 查询字段 (下拉框，可编辑)
        qfield_combo = QComboBox()
        qfield_combo.setEditable(True)
        qfield_combo.addItem(_t("（无）"), "")
        fields = self._get_module_fields()
        for f in fields:
            qfield_combo.addItem(f, f)

        if qfield:
            idx = qfield_combo.findData(qfield)
            if idx >= 0:
                qfield_combo.setCurrentIndex(idx)
            else:
                qfield_combo.addItem(qfield, qfield)
                qfield_combo.setCurrentIndex(qfield_combo.count() - 1)
        else:
            qfield_combo.setCurrentIndex(0)
        tbl.setCellWidget(r, 0, qfield_combo)

        # 列 1: 组件 ID (下拉框)
        combo = QComboBox()
        combo.addItem("（无）", "")
        # 获取当前交互模块的所有组件 ID 并添加入下拉框中
        comp_ids = self._get_im_component_ids()
        for cid in comp_ids:
            combo.addItem(cid, cid)

        # 预设值
        if ui_comp:
            idx = combo.findData(ui_comp)
            if idx >= 0:
                combo.setCurrentIndex(idx)
            else:
                # 兼容性：如果原来的值不在列表中（例如用户改过交互模块），把它临时加进下拉框以免丢失
                combo.addItem(ui_comp, ui_comp)
                combo.setCurrentIndex(combo.count() - 1)
        tbl.setCellWidget(r, 1, combo)

        # 列 2: Excel 列 (文本输入)
        tbl.setItem(r, 2, QTableWidgetItem(cols))

        # 列 3: 拼接符 (文本输入)
        tbl.setItem(r, 3, QTableWidgetItem(sep))

        # 列 4: 查询默认值 (文本输入)
        tbl.setItem(r, 4, QTableWidgetItem(def_val))

    def _on_im_changed(self) -> None:
        """当交互模块变更时，动态刷新输入映射表中所有组件ID的下拉框列表"""
        comp_ids = self._get_im_component_ids()
        for r in range(self._in_table.rowCount()):
            w = self._in_table.cellWidget(r, 1)
            if isinstance(w, QComboBox):
                cur_val = w.currentData() or w.currentText()
                w.blockSignals(True)
                w.clear()
                w.addItem("（无）", "")
                for cid in comp_ids:
                    w.addItem(cid, cid)

                # 重新定位先前选中的值
                idx = w.findData(cur_val)
                if idx >= 0:
                    w.setCurrentIndex(idx)
                elif cur_val:
                    w.addItem(cur_val, cur_val)
                    w.setCurrentIndex(w.count() - 1)
                w.blockSignals(False)

    def _field_combo_row(self, tbl: QTableWidget, field: str = "") -> QComboBox:
        """在 tbl 最后一行的第 0 列插入字段 ComboBox，返回 combo"""
        r = tbl.rowCount(); tbl.insertRow(r)
        combo = QComboBox()
        for f in self._get_module_fields():
            combo.addItem(f)
        if field:
            idx = combo.findText(field)
            if idx >= 0:
                combo.setCurrentIndex(idx)
            else:
                combo.addItem(field); combo.setCurrentIndex(combo.count() - 1)
        tbl.setCellWidget(r, 0, combo)
        return combo

    def _add_fields_to_disp(self, fields: list[str]) -> None:
        """批量添加字段到显示列（字段名直接放 col 0，可编辑）"""
        for f in fields:
            r = self._disp_table.rowCount()
            self._disp_table.insertRow(r)
            self._disp_table.setItem(r, 0, QTableWidgetItem(f))
            self._disp_table.setItem(r, 1, QTableWidgetItem(f))
            self._disp_table.setItem(r, 2, QTableWidgetItem("120"))

    def _pick_output_fields(self, tbl: QTableWidget) -> None:
        fields = self._get_module_fields()
        if not fields:
            self._field_combo_row(tbl, "")
            tbl.setItem(tbl.rowCount() - 1, 1, QTableWidgetItem(""))
            return
        dlg = _FieldPickerDialog(fields, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            sel = dlg.selected_fields()
        else:
            sel = [""]
        for f in (sel or [""]):
            self._field_combo_row(tbl, f)
            tbl.setItem(tbl.rowCount() - 1, 1, QTableWidgetItem(""))

    def _del_row(self, tbl: QTableWidget) -> None:
        r = tbl.currentRow()
        if r >= 0:
            tbl.removeRow(r)

    def _move_row(self, tbl: QTableWidget, d: int) -> None:
        r = tbl.currentRow(); t = r + d
        if r < 0 or not (0 <= t < tbl.rowCount()):
            return
        for c in range(tbl.columnCount()):
            wa = tbl.cellWidget(r, c); wb = tbl.cellWidget(t, c)
            ia = tbl.takeItem(r, c);   ib = tbl.takeItem(t, c)
            if wa or wb:
                ta = wa.currentText() if isinstance(wa, QComboBox) else ""
                tb = wb.currentText() if isinstance(wb, QComboBox) else ""
                def _rebuild(src, txt, row, col):
                    if isinstance(src, QComboBox):
                        nb = QComboBox()
                        nb.setEditable(src.isEditable())
                        for i in range(src.count()):
                            nb.addItem(src.itemText(i), src.itemData(i))
                        idx = nb.findText(txt)
                        if idx >= 0:
                            nb.setCurrentIndex(idx)
                        elif txt:
                            nb.setEditText(txt)
                        tbl.setCellWidget(row, col, nb)
                _rebuild(wa, ta, t, c)
                _rebuild(wb, tb, r, c)
            else:
                if ib: tbl.setItem(r, c, ib)
                if ia: tbl.setItem(t, c, ia)
        tbl.setCurrentCell(t, tbl.currentColumn())

    # ── 数据填充 ─────────────────────────────────────────────────

    def _load_group(self, group) -> None:
        for im in group.input_mappings:
            self._in_mapping_row(
                self._in_table,
                im.query_field,
                im.ui_component,
                ",".join(im.columns),
                im.separator,
                getattr(im, "default_value", ""),
            )
        for dc in group.display_columns:
            r = self._disp_table.rowCount(); self._disp_table.insertRow(r)
            self._disp_table.setItem(r, 0, QTableWidgetItem(dc.field))
            self._disp_table.setItem(r, 1, QTableWidgetItem(dc.header))
            self._disp_table.setItem(r, 2, QTableWidgetItem(str(dc.width)))
        for om in group.output_mappings:
            self._field_combo_row(self._out_table, om.result_field)
            self._out_table.setItem(self._out_table.rowCount() - 1, 1, QTableWidgetItem(om.column))

    # ── 读取表格数据 ─────────────────────────────────────────────

    def _cell(self, tbl: QTableWidget, r: int, c: int) -> str:
        w = tbl.cellWidget(r, c)
        if isinstance(w, QComboBox):
            val = w.currentData()
            if val is not None:
                return str(val).strip()
            return w.currentText().strip()
        it = tbl.item(r, c)
        return it.text().strip() if it else ""

    # ── 保存 ─────────────────────────────────────────────────────

    def _on_ok(self) -> None:
        from core.mapping import MappingGroup, InputMapping, OutputMapping, DisplayColumn

        name = self._name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "错误", "映射组名称不能为空"); return

        gid = self._group_id if self._group_id else ConfigManager.generate_id("mg_")

        input_mappings = []
        for r in range(self._in_table.rowCount()):
            qf  = self._cell(self._in_table, r, 0)
            uic = self._cell(self._in_table, r, 1)  # UI组件 field
            def_val = self._cell(self._in_table, r, 4)
            if not qf and not uic and not def_val: continue

            cols = [c.strip() for c in self._cell(self._in_table, r, 2).split(",") if c.strip()]
            sep  = self._cell(self._in_table, r, 3) or " "
            input_mappings.append(InputMapping(
                query_field=qf, ui_component=uic, columns=cols, separator=sep, default_value=def_val
            ))

        display_columns = []
        for r in range(self._disp_table.rowCount()):
            f = self._cell(self._disp_table, r, 0)
            if not f: continue
            hdr = self._cell(self._disp_table, r, 1) or f
            w   = self._cell(self._disp_table, r, 2)
            display_columns.append(DisplayColumn(field=f, header=hdr, width=int(w) if w.isdigit() else 120))

        output_mappings = []
        for r in range(self._out_table.rowCount()):
            f = self._cell(self._out_table, r, 0)
            c = self._cell(self._out_table, r, 1)
            if f and c:
                output_mappings.append(OutputMapping(result_field=f, column=c))

        self._result_id = gid
        self._result_group = MappingGroup(
            name=name,
            interaction_module=self._im_combo.currentData() or "",
            query_module=self._module_combo.currentData(),
            input_mappings=input_mappings,
            output_mappings=output_mappings,
            display_columns=display_columns,
            query_params={"limit": self._limit_spin.value()},
        )
        self.accept()

    def get_result(self):
        return self._result_id, self._result_group


# ══════════════════════════════════════════════════════════════════
# Tab 5: 交互模块管理
# ══════════════════════════════════════════════════════════════════

class _InteractionModuleTab(QWidget):
    """
    交互模块管理标签页。 / Interaction Module Management Tab.
    左右分栏：左侧列表 + 右侧预览。 / Split Layout: Left list + Right live preview.
    """

    def __init__(self, app_controller: Application):
        super().__init__()
        self._app = app_controller
        layout = QHBoxLayout(self)

        # 左侧：交互模块列表
        left = QVBoxLayout()
        left.addWidget(QLabel(_t("交互模块列表:")))
        self._list = QListWidget()
        self._list.currentRowChanged.connect(self._on_select)
        left.addWidget(self._list)

        btn_bar = QHBoxLayout()
        add_btn = QPushButton("➕ 新建")
        add_btn.clicked.connect(self._add_module)
        edit_btn = QPushButton("✏ 编辑")
        edit_btn.clicked.connect(self._edit_module)
        del_btn = QPushButton("🗑 删除")
        del_btn.clicked.connect(self._delete_module)
        btn_bar.addWidget(add_btn)
        btn_bar.addWidget(edit_btn)
        btn_bar.addWidget(del_btn)
        left.addLayout(btn_bar)

        # 右侧：预览
        right = QVBoxLayout()
        right.addWidget(QLabel(_t("交互模块详情:")))
        self._preview = QLabel(_t("（请选择交互模块）"))
        self._preview.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        self._preview.setWordWrap(True)
        self._preview.setStyleSheet("background: #f5f5f5; padding: 10px; border-radius: 4px;")
        right.addWidget(self._preview, 1)

        layout.addLayout(left, 1)
        layout.addLayout(right, 2)

    def refresh(self) -> None:
        self._list.clear()
        ims = self._app.config_mgr.get_interaction_modules()
        for im_id, im in ims.items():
            bt_label = _t(BASE_TYPE_LABELS.get(im.base_type, im.base_type))
            item = QListWidgetItem(f"{im.name} [{bt_label}] [{im_id}]")
            item.setData(Qt.UserRole, im_id)
            self._list.addItem(item)
        self._preview.setText(_t("（请选择交互模块）"))

    def _on_select(self, row: int) -> None:
        if row < 0:
            self._preview.setText(_t("（请选择交互模块）"))
            return
        im_id = self._list.item(row).data(Qt.UserRole)
        im = self._app.config_mgr.get_interaction_module(im_id)
        if im is None:
            return
        bt_label = _t(BASE_TYPE_LABELS.get(im.base_type, im.base_type))
        lines = [
            f"<b>{_t('名称')}:</b> {im.name}",
            f"<b>{_t('基础类型')}:</b> {bt_label} ({im.base_type})",
        ]
        if im.base_type == "display_panel" and im.components:
            lines.append("")
            lines.append(f"<b>{_t('组件列表')}:</b>")
            for c in im.components:
                ct_label = _t(COMP_TYPE_LABELS.get(c.comp_type, c.comp_type))
                lines.append(f"  • {c.label} [{ct_label}] → {c.field}")
        elif im.base_type == "popup_table" and im.ui_controls:
            lines.append("")
            lines.append(f"<b>{_t('UI 控件')}:</b>")
            for uc in im.ui_controls:
                if uc.control_type == "checkbox":
                    lines.append(f"  • {_t('复选框')} [{uc.label}]")
                else:
                    opts = ", ".join(o.get("label", "") for o in uc.options)
                    lines.append(f"  • {_t('下拉')} [{uc.label}]  {_t('选项')}: {opts}")
        elif im.base_type == "search_panel" and im.search_boxes:
            lines.append("")
            lines.append(f"<b>{_t('搜索框')}:</b>")
            for sb in im.search_boxes:
                lines.append(f"  • {sb.get('label', '')} → {sb.get('field', '')}  {_t('宽')} {sb.get('width', 140)}px")
        elif im.base_type == "auto_fill":
            lines.append(f"<b>{_t('自动关闭')}:</b> {im.auto_close_seconds} {_t('秒')}")
        self._preview.setText("<br>".join(lines))


    def _add_module(self) -> None:
        dlg = _InteractionModuleDialog(self._app, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            im_id, im = dlg.get_result()
            self._app.config_mgr.save_interaction_module(im_id, im)
            self.refresh()

    def _edit_module(self) -> None:
        row = self._list.currentRow()
        if row < 0:
            QMessageBox.information(self, _t("提示"), _t("请先选择要编辑的交互模块"))
            return
        im_id = self._list.item(row).data(Qt.UserRole)
        im = self._app.config_mgr.get_interaction_module(im_id)
        if im is None:
            return
        dlg = _InteractionModuleDialog(self._app, im_id=im_id, im=im, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            new_id, new_im = dlg.get_result()
            if new_id != im_id:
                self._app.config_mgr.delete_interaction_module(im_id)
            self._app.config_mgr.save_interaction_module(new_id, new_im)
            self.refresh()

    def _delete_module(self) -> None:
        row = self._list.currentRow()
        if row < 0:
            return
        im_id = self._list.item(row).data(Qt.UserRole)
        im = self._app.config_mgr.get_interaction_module(im_id)
        display = f"{im.name} [{im_id}]" if im else im_id
        # 检查引用
        refs = self._app.config_mgr.get_referencing_mg_for_im(im_id)
        if refs:
            ref_list = "\n".join(f"  • {r}" for r in refs)
            ans = QMessageBox.warning(
                self, _t("确认删除（存在引用）"),
                _t("以下映射组引用了交互模块「{display}」：\n{ref_list}\n\n"
                "删除后这些映射组的交互模块字段将被清空，确认删除？"),
                QMessageBox.Yes | QMessageBox.No,
            )
        else:
            ans = QMessageBox.question(self, _t("确认删除"), _t("确定删除交互模块「{display}」？").format(display=display))
        if ans == QMessageBox.Yes:
            self._app.config_mgr.delete_interaction_module(im_id)
            self.refresh()


# ══════════════════════════════════════════════════════════════════
# 交互模块编辑对话框
# ══════════════════════════════════════════════════════════════════

class _InteractionModuleDialog(QDialog):
    """创建或编辑交互模块的弹窗"""

    def __init__(self, app_controller: Application, im_id: str = "", im: InteractionModule | None = None, parent=None):
        super().__init__(parent)
        self._app = app_controller
        self._im_id = im_id  # 编辑时保留；新建时为空
        self._editing = bool(im_id)
        self.setWindowTitle(_t("编辑交互模块") if im_id else _t("新建交互模块"))
        self.setMinimumSize(700, 500)
        self._setup_ui(im)

    def _setup_ui(self, im: InteractionModule | None) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 8)
        layout.setSpacing(8)

        # ── 基本信息 ─────────────────────────────────────────────
        basic = QGroupBox(_t("基本信息"))
        bf = QFormLayout(basic)
        bf.setFieldGrowthPolicy(QFormLayout.ExpandingFieldsGrow)

        self._name_input = QLineEdit(im.name if im else "")
        self._name_input.setPlaceholderText(_t("交互模块名称"))
        bf.addRow(_t("名称:"), self._name_input)

        self._type_combo = QComboBox()
        for bt in BASE_TYPES:
            self._type_combo.addItem(f"{_t(BASE_TYPE_LABELS.get(bt, bt))} ({bt})", bt)
        if im:
            idx = self._type_combo.findData(im.base_type)
            if idx >= 0:
                self._type_combo.setCurrentIndex(idx)
        # 编辑时基础类型只读
        if self._editing:
            self._type_combo.setEnabled(False)
        self._type_combo.currentIndexChanged.connect(self._on_type_changed)
        bf.addRow(_t("基础类型:"), self._type_combo)

        layout.addWidget(basic)

        # ── 配置区域容器（根据 base_type 切换） ───────────────────
        self._config_stack = QWidget()
        self._config_layout = QVBoxLayout(self._config_stack)
        self._config_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self._config_stack, 1)

        # 预建各类型配置面板
        self._dp_widget = self._build_display_panel_config(im)
        self._sp_widget = self._build_search_panel_config(im)
        self._pt_widget = self._build_popup_table_config(im)
        self._af_widget = self._build_auto_fill_config(im)

        self._config_layout.addWidget(self._dp_widget)
        self._config_layout.addWidget(self._sp_widget)
        self._config_layout.addWidget(self._pt_widget)
        self._config_layout.addWidget(self._af_widget)

        self._on_type_changed()

        # ── 底部按钮 ─────────────────────────────────────────────
        bar = QHBoxLayout()
        bar.addStretch()
        ok = QPushButton(_t("✔ 确定"))
        ok.setFixedWidth(90)
        ok.clicked.connect(self._on_ok)
        ca = QPushButton(_t("✘ 取消"))
        ca.setFixedWidth(90)
        ca.clicked.connect(self.reject)
        bar.addWidget(ok); bar.addWidget(ca)
        layout.addLayout(bar)

    def _on_type_changed(self) -> None:
        """根据选中的基础类型显示对应的配置面板"""
        bt = self._type_combo.currentData() or "display_panel"
        self._dp_widget.setVisible(bt == "display_panel")
        self._sp_widget.setVisible(bt == "search_panel")
        self._pt_widget.setVisible(bt == "popup_table")
        self._af_widget.setVisible(bt == "auto_fill")

    # ── display_panel 组件编辑 ────────────────────────────────────

    def _build_display_panel_config(self, im: InteractionModule | None) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.addWidget(QLabel(_t("组件列表（定义自定义面板中的 UI 组件）:")))

        self._comp_table = QTableWidget(0, 5)
        self._comp_table.setHorizontalHeaderLabels([_t("标签"), _t("类型"), _t("组件ID"), _t("下拉选项"), _t("默认值")])
        hdr = self._comp_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.Fixed)
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)
        hdr.setSectionResizeMode(3, QHeaderView.Stretch)
        hdr.setSectionResizeMode(4, QHeaderView.Fixed)
        self._comp_table.setColumnWidth(1, 110)
        self._comp_table.setColumnWidth(4, 90)
        self._comp_table.setAlternatingRowColors(True)
        v.addWidget(self._comp_table)

        btn_bar = QHBoxLayout()
        add_btn = QPushButton(_t("➕ 添加组件"))
        add_btn.clicked.connect(self._add_comp_row)
        del_btn = QPushButton(_t("🗑 删除行"))
        del_btn.clicked.connect(lambda: self._del_table_row(self._comp_table))
        up_btn = QPushButton("↑"); up_btn.setFixedWidth(30)
        up_btn.clicked.connect(lambda: self._move_table_row(self._comp_table, -1))
        dn_btn = QPushButton("↓"); dn_btn.setFixedWidth(30)
        dn_btn.clicked.connect(lambda: self._move_table_row(self._comp_table, +1))
        btn_bar.addWidget(add_btn)
        btn_bar.addWidget(del_btn)
        btn_bar.addSpacing(8)
        btn_bar.addWidget(up_btn)
        btn_bar.addWidget(dn_btn)
        btn_bar.addStretch()
        v.addLayout(btn_bar)

        # 填充已有数据
        if im and im.base_type == "display_panel":
            for c in im.components:
                self._add_comp_row(c)

        return w

    def _add_comp_row(self, comp: ComponentDef | None = None) -> None:
        """添加一行组件定义"""
        row = self._comp_table.rowCount()
        self._comp_table.insertRow(row)

        # 列 0: 标签
        self._comp_table.setItem(row, 0, QTableWidgetItem(comp.label if comp else ""))

        # 列 1: 类型下拉
        tc = QComboBox()
        for ct in ALL_COMP_TYPES:
            tc.addItem(_t(COMP_TYPE_LABELS.get(ct, ct)), ct)
        if comp:
            idx = tc.findData(comp.comp_type)
            if idx >= 0:
                tc.setCurrentIndex(idx)
        self._comp_table.setCellWidget(row, 1, tc)

        # 列 2: 查询字段
        self._comp_table.setItem(row, 2, QTableWidgetItem(comp.field if comp else ""))

        # 列 3: 下拉选项（逗号分隔）
        opts_str = ",".join(comp.options) if comp and comp.options else ""
        self._comp_table.setItem(row, 3, QTableWidgetItem(opts_str))

        # 列 4: 默认值
        self._comp_table.setItem(row, 4, QTableWidgetItem(comp.default if comp else ""))

    # ── search_panel 搜索框编辑 ──────────────────────────────────

    def _build_search_panel_config(self, im: InteractionModule | None) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.addWidget(QLabel(_t("搜索框列表（定义搜索面板中的输入框）:")))

        self._sb_table = QTableWidget(0, 3)
        self._sb_table.setHorizontalHeaderLabels([_t("组件ID"), _t("标签"), _t("宽度")])
        hdr = self._sb_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        self._sb_table.setColumnWidth(2, 70)
        self._sb_table.setAlternatingRowColors(True)
        v.addWidget(self._sb_table)

        btn_bar = QHBoxLayout()
        add_btn = QPushButton(_t("➕ 添加搜索框"))
        add_btn.clicked.connect(lambda: self._add_sb_row())
        del_btn = QPushButton(_t("🗑 删除行"))
        del_btn.clicked.connect(lambda: self._del_table_row(self._sb_table))
        btn_bar.addWidget(add_btn)
        btn_bar.addWidget(del_btn)
        btn_bar.addStretch()
        v.addLayout(btn_bar)

        # 填充已有数据
        if im and im.base_type == "search_panel":
            for sb in im.search_boxes:
                self._add_sb_row(sb)

        return w

    def _add_sb_row(self, sb: dict | None = None) -> None:
        row = self._sb_table.rowCount()
        self._sb_table.insertRow(row)
        self._sb_table.setItem(row, 0, QTableWidgetItem(sb.get("field", "") if sb else ""))
        self._sb_table.setItem(row, 1, QTableWidgetItem(sb.get("label", "") if sb else ""))
        self._sb_table.setItem(row, 2, QTableWidgetItem(str(sb.get("width", 140)) if sb else "140"))

    # ── popup_table UI 控件编辑 ──────────────────────────────────

    def _build_popup_table_config(self, im: InteractionModule | None) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        tip = QLabel(
            _t("定义弹窗查询 of UI 过滤控件（复选框/下拉）。\n"
            "控件的查询字段需在映射组「📥 输入映射」中以「UI组件模式」关联。\n"
            "复选框: 选中/未选中值填 True/False 或字符串，留空=不传该字段\n"
            "下拉框: 选项格式 标签1=值1;标签2=值2，值为空=不传该字段")
        )
        tip.setStyleSheet("color:#666;font-size:11px")
        v.addWidget(tip)

        self._ui_table = QTableWidget(0, 6)
        self._ui_table.setHorizontalHeaderLabels(
            [_t("类型"), _t("标签"), _t("选中值"), _t("未选中值"), _t("默认选中/索引"), _t("下拉选项")]
        )
        hdr = self._ui_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr.setSectionResizeMode(5, QHeaderView.Stretch)
        for i in range(1, 5):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self._ui_table.setColumnWidth(0, 80)
        self._ui_table.setAlternatingRowColors(True)
        v.addWidget(self._ui_table)

        btn_bar = QHBoxLayout()
        cb_btn = QPushButton(_t("➕ 复选框"))
        cb_btn.clicked.connect(lambda: self._add_ui_row(
            ["checkbox", "仅查有效", "True", "False", "是", ""]))
        dd_btn = QPushButton(_t("➕ 下拉框"))
        dd_btn.clicked.connect(lambda: self._add_ui_row(
            ["dropdown", _t("状态"), "", "", "0", _t("全部=;有效=有效;无效=无效")]))
        del_btn = QPushButton(_t("🗑 删除行"))
        del_btn.clicked.connect(lambda: self._del_table_row(self._ui_table))
        btn_bar.addWidget(cb_btn)
        btn_bar.addWidget(dd_btn)
        btn_bar.addWidget(del_btn)
        btn_bar.addStretch()
        v.addLayout(btn_bar)

        # 填充已有数据
        if im and im.base_type == "popup_table":
            for uc in im.ui_controls:
                if uc.control_type == "checkbox":
                    cv = "" if uc.checked_value is None else str(uc.checked_value)
                    uv = "" if uc.unchecked_value is None else str(uc.unchecked_value)
                    self._add_ui_row(["checkbox", uc.label, cv, uv,
                                      "是" if uc.default_checked else "否", ""])
                elif uc.control_type == "dropdown":
                    opts = ";".join(f"{o['label']}={o['value']}" for o in uc.options)
                    self._add_ui_row(["dropdown", uc.label, "", "",
                                      str(uc.default_index), opts])

        return w

    def _add_ui_row(self, vals: list[str]) -> None:
        row = self._ui_table.rowCount()
        self._ui_table.insertRow(row)
        for c, v in enumerate(vals):
            self._ui_table.setItem(row, c, QTableWidgetItem(v))

    # ── auto_fill 配置 ───────────────────────────────────────────

    def _build_auto_fill_config(self, im: InteractionModule | None) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.addWidget(QLabel(_t("自动填充模式配置:")))

        f = QFormLayout()
        self._auto_close_spin = QSpinBox()
        self._auto_close_spin.setRange(0, 300)
        self._auto_close_spin.setSuffix(_t(" 秒"))
        self._auto_close_spin.setValue(im.auto_close_seconds if im else 3)
        f.addRow(_t("自动关闭时间:"), self._auto_close_spin)
        v.addLayout(f)
        v.addStretch()

        return w

    # ── 通用表格辅助 ─────────────────────────────────────────────

    def _del_table_row(self, tbl: QTableWidget) -> None:
        r = tbl.currentRow()
        if r >= 0:
            tbl.removeRow(r)

    def _move_table_row(self, tbl: QTableWidget, d: int) -> None:
        r = tbl.currentRow(); t = r + d
        if r < 0 or not (0 <= t < tbl.rowCount()):
            return
        for c in range(tbl.columnCount()):
            wa = tbl.cellWidget(r, c); wb = tbl.cellWidget(t, c)
            ia = tbl.takeItem(r, c);   ib = tbl.takeItem(t, c)
            if wa or wb:
                ta = wa.currentText() if isinstance(wa, QComboBox) else ""
                tb = wb.currentText() if isinstance(wb, QComboBox) else ""
                def _rebuild(src, txt, row_, col_):
                    if isinstance(src, QComboBox):
                        nb = QComboBox()
                        for i in range(src.count()): nb.addItem(src.itemText(i))
                        idx = nb.findText(txt)
                        if idx >= 0: nb.setCurrentIndex(idx)
                        tbl.setCellWidget(row_, col_, nb)
                _rebuild(wa, ta, t, c)
                _rebuild(wb, tb, r, c)
            else:
                if ib: tbl.setItem(r, c, ib)
                if ia: tbl.setItem(t, c, ia)
        tbl.setCurrentCell(t, tbl.currentColumn())

    def _cell(self, tbl: QTableWidget, r: int, c: int) -> str:
        """读取表格单元格文本（兼容 QComboBox 和 QTableWidgetItem）"""
        w = tbl.cellWidget(r, c)
        if isinstance(w, QComboBox):
            return w.currentData() or w.currentText().strip()
        it = tbl.item(r, c)
        return it.text().strip() if it else ""

    # ── 保存 ─────────────────────────────────────────────────────

    def _on_ok(self) -> None:
        from core.mapping import UIControlMapping

        name = self._name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "错误", "交互模块名称不能为空"); return

        bt = self._type_combo.currentData() or "display_panel"
        im_id = self._im_id if self._im_id else ConfigManager.generate_id("im_")

        components: list[ComponentDef] = []
        ui_controls: list[UIControlMapping] = []
        search_boxes: list[dict] = []
        auto_close_seconds = 3

        if bt == "display_panel":
            for r in range(self._comp_table.rowCount()):
                label = self._cell(self._comp_table, r, 0)
                if not label:
                    continue
                comp_type = self._cell(self._comp_table, r, 1)
                field = self._cell(self._comp_table, r, 2)
                opts_str = self._cell(self._comp_table, r, 3)
                options = [o.strip() for o in opts_str.split(",") if o.strip()] if opts_str else []
                default = self._cell(self._comp_table, r, 4)
                components.append(ComponentDef(
                    label=label, comp_type=comp_type, field=field,
                    options=options, default=default,
                ))

        elif bt == "search_panel":
            for r in range(self._sb_table.rowCount()):
                field = self._cell(self._sb_table, r, 0)
                if not field:
                    continue
                label = self._cell(self._sb_table, r, 1) or field
                w_str = self._cell(self._sb_table, r, 2)
                width = int(w_str) if w_str.isdigit() else 140
                search_boxes.append({"field": field, "label": label, "width": width})

        elif bt == "popup_table":
            for r in range(self._ui_table.rowCount()):
                ct    = self._cell(self._ui_table, r, 0).lower()
                label = self._cell(self._ui_table, r, 1)
                if not label:
                    continue
                # query_field 自动取 label，作为控件标识符
                # InputMapping.ui_component 填写该标签即可关联
                def _pv(s):
                    if s == "": return None
                    if s.lower() == "true": return True
                    if s.lower() == "false": return False
                    return s
                if ct == "checkbox":
                    ui_controls.append(UIControlMapping(
                        control_type="checkbox",
                        label=label,
                        query_field=label,          # 标识符 = 标签
                        checked_value=_pv(self._cell(self._ui_table, r, 2)),
                        unchecked_value=_pv(self._cell(self._ui_table, r, 3)),
                        default_checked=self._cell(self._ui_table, r, 4).strip() not in ("否", "0", "false", "False"),
                    ))
                elif ct == "dropdown":
                    opts = []
                    for part in self._cell(self._ui_table, r, 5).split(";"):
                        if "=" in part:
                            lb, vl = part.split("=", 1)
                            opts.append({"label": lb.strip(), "value": vl.strip()})
                    di = self._cell(self._ui_table, r, 4)
                    ui_controls.append(UIControlMapping(
                        control_type="dropdown",
                        label=label,
                        query_field=label,          # 标识符 = 标签
                        options=opts,
                        default_index=int(di) if di.isdigit() else 0,
                    ))

        elif bt == "auto_fill":
            auto_close_seconds = self._auto_close_spin.value()

        self._result_id = im_id
        self._result_im = InteractionModule(
            name=name,
            base_type=bt,
            components=components,
            ui_controls=ui_controls,
            search_boxes=search_boxes,
            auto_close_seconds=auto_close_seconds,
        )
        self.accept()

    def get_result(self) -> tuple[str, InteractionModule]:
        return self._result_id, self._result_im

