"""
test/generate_test_data.py

Automatically generates large-scale realistic test data for the Excel Helper.
Generates:
  1. test_inventory.xlsx (10,500 rows of spare parts, including a random '出厂批次' field)
  2. test_database.db (SQLite database matching the Excel structure)
  3. test_interactive_query.xlsx (User interactive Excel file with blank fields, bold styling, column widths, and instructions)
"""

import os
import sqlite3
import random

def make_realistic_data(count=10500):
    """Generates deterministic, highly realistic spare parts records with a random batch number."""
    prefixes = ["高强度", "不锈钢", "进口", "防爆", "耐高温", "精密", "耐磨", "特高压", "合金", "超轻"]
    nouns = ["轴承", "螺栓", "螺母", "阀门", "联轴器", "齿轮", "油泵", "密封圈", "传感器", "电容", "继电器", "液压缸"]
    categories = ["A型", "B型", "双金属", "重型", "微型", "气动", "数显", "高精度", "直角", "直通"]
    specs_template = ["620{i}", "M{i}x50", "DN{i}", "YE{i}", "HL{i}", "Q41F-{i}", "DN{i} L={len}m", "YE3-{i} 5.5kW"]
    units = ["个", "套", "台", "根", "米"]
    
    # Using a fixed seed for deterministic, reproducible records
    random.seed(42)
    
    data = []
    for idx in range(count):
        code = f"M{10001 + idx}"
        
        pref = random.choice(prefixes)
        noun = random.choice(nouns)
        cat = random.choice(categories)
        name = f"{pref}{cat}{noun}"
        
        tmpl = random.choice(specs_template)
        num = random.randint(10, 999)
        length = random.choice([3, 6, 9])
        spec = tmpl.format(i=num, len=length)
        
        unit = random.choice(units)
        stock = random.randint(10, 10000)
        price = round(random.uniform(0.50, 8500.00), 2)
        
        # ── 新增随机生成字段: "出厂批次" ─────────────────────────────────
        batch_num = random.randint(10000, 99999)
        batch = f"BAT-{batch_num}"
        
        data.append([code, name, spec, unit, stock, price, batch])
    return data

def generate_excel(path, rows_data):
    """Generates a large-scale Excel file with spare parts data."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("[-] Error: openpyxl is not installed. Please run 'pip install openpyxl' first.")
        return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "备件库"
    
    # Headers containing '出厂批次'
    headers = ["物料代码", "物资名称", "规格型号", "单位", "库存", "单价", "出厂批次"]
    ws.append(headers)
    
    for row in rows_data:
        ws.append(row)
        
    wb.save(path)
    print(f"[+] Excel test file generated successfully with {len(rows_data)} rows at: {path}")

def generate_sqlite(path, rows_data):
    """Generates a matching SQLite database file with large-scale data."""
    # Delete first if already exists to ensure fresh generation
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError as e:
            print(f"[-] Warning: could not remove existing db file: {e}")
            
    conn = sqlite3.connect(path)
    cursor = conn.cursor()
    
    # Create Table including factory_batch
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS materials (
        material_code TEXT PRIMARY KEY,
        material_name TEXT,
        spec TEXT,
        unit TEXT,
        stock INTEGER,
        price REAL,
        factory_batch TEXT
    )
    """)
    
    cursor.executemany("""
    INSERT INTO materials (material_code, material_name, spec, unit, stock, price, factory_batch)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, rows_data)
    
    conn.commit()
    conn.close()
    print(f"[+] SQLite test database generated successfully with {len(rows_data)} rows at: {path}")

def generate_interactive_excel(path, source_data):
    """Generates a styled, rich interactive test Excel workbook with dual sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[-] Error: openpyxl is not installed. Skipping styled generation.")
        return
        
    wb = Workbook()
    
    # ── Sheet 1: 查询看板 ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "查询看板"
    
    # Stylings
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    headers = ["物料代码", "物资名称", "规格型号", "单位", "库存", "单价", "出厂批次"]
    ws1.append(headers)
    
    # Style header row
    for col_idx in range(1, 8):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Pick 70 codes (leave name blank) and 30 names (leave code blank)
    random.seed(99)
    sample_rows = random.sample(source_data, 100)
    
    # Populate data
    for idx, item in enumerate(sample_rows):
        row_num = idx + 2
        code, name, spec, unit, stock, price, batch = item
        
        if idx < 70:
            # Code-driven: only code is prefilled, rest are blank
            row_data = [code, "", "", "", "", "", ""]
        else:
            # Name-driven: only name is prefilled, rest are blank
            row_data = ["", name, "", "", "", "", ""]
            
        ws1.append(row_data)
        
        # Style data row
        for col_idx in range(1, 8):
            cell = ws1.cell(row=row_num, column=col_idx)
            cell.font = data_font
            if col_idx in (1, 4, 5, 6, 7):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Column widths
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 16
    
    # ── Sheet 2: 使用说明 ───────────────────────────────────────────
    ws2 = wb.create_sheet(title="测试与使用说明")
    ws2.column_dimensions["A"].width = 65
    
    title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E78")
    section_font = Font(name="微软雅黑", size=11, bold=True)
    body_font = Font(name="微软雅黑", size=10)
    
    ws2.cell(row=1, column=1, value="Excel 快捷查询助手 — 测试沙盒说明书").font = title_font
    
    instructions = [
        "",
        "【测试准备】",
        "1. 启动 Excel 快捷查询助手并确保托盘常驻。",
        "2. 系统已自动应用 `test/` 相对路径配置，无需任何手动设定即可开始。",
        "",
        "【玩法 A: 单行双击写回 (双向支持)】",
        "1. 切换至「查询看板」工作表。",
        "2. 选中带有“物料代码”（A列）或“物资名称”（B列）的任意单元格。",
        "3. 按下快捷键：",
        "   - [Excel 文件查询]：按下 Ctrl + Shift + F",
        "   - [SQLite 数据库查询]：按下 Ctrl + Shift + D",
        "4. 自定义交互面板将秒级展示，点击「查询」按钮。",
        "5. 在搜索结果表格中，双击任意一行，该行数据将秒级填充回本表的空白单元格中！",
        "",
        "【玩法 B: 多行智能批量自动填充 (Auto Fill)】",
        "1. 用鼠标框选「查询看板」中的多行（如第 2 到 15 行，支持Ctrl多选不连续区域）。",
        "2. 按下自动填充的触发命令（或在托盘上右击并选择运行对应模块）。",
        "3. 极速进度条对话框将启动，一键在后台批量完成所有查询与自动写回！",
    ]
    
    for r_idx, text in enumerate(instructions):
        cell = ws2.cell(row=r_idx + 2, column=1, value=text)
        if text.startswith("【"):
            cell.font = section_font
        else:
            cell.font = body_font
            
    wb.save(path)
    print(f"[+] Rich interactive query Excel test file generated successfully at: {path}")

if __name__ == "__main__":
    dir_path = os.path.dirname(os.path.abspath(__file__))
    
    excel_path = os.path.join(dir_path, "test_inventory.xlsx")
    sqlite_path = os.path.join(dir_path, "test_database.db")
    interactive_path = os.path.join(dir_path, "test_interactive_query.xlsx")
    
    print("=" * 60)
    print("  Excel Helper - Generating Test Environment & Rich Interactive Sheet")
    print("=" * 60)
    
    # Generate 10,500 rows of matched records
    test_rows = make_realistic_data(10500)
    
    generate_excel(excel_path, test_rows)
    generate_sqlite(sqlite_path, test_rows)
    generate_interactive_excel(interactive_path, test_rows)
    
    print("-" * 60)
    print("Success: Large-scale test dataset and interactive workbook generated successfully!")
    print("You can now test query write-back and Auto-Fill in:")
    print(f"  - Interactive Workbook: {interactive_path}")
    print(f"  - Excel Source:         {excel_path} (Sheet: '备件库')")
    print(f"  - SQLite Database:      {sqlite_path} (Table: 'materials')")
    print("=" * 60)
