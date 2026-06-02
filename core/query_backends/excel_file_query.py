"""
core/query_backends/excel_file_query.py

Excel 文件查询后端 —— 从本地 Excel 文件读取数据，在内存中进行检索。

读取引擎（按优先级自动选择）：
  1. python-calamine  Rust 实现，推荐，速度是 openpyxl 的 10-30 倍
                      同时支持 .xlsx/.xlsm/.xls/.ods 等格式
  2. openpyxl         纯 Python 兜底，仅支持 .xlsx/.xlsm
  两者均未安装时 connect() 返回 False 并报错。

功能特性：
  - 单个或多个 Excel 文件 / 遍历目录（可选递归）作为数据源
  - 首行表头模式 / 无表头模式
  - 按列字母（A/B/C）、1-based 序号（1/2/3）或表头字段名配置列映射
  - 留空列配置 = 自动读取全部列
  - ILIKE 模糊匹配（默认）与正则匹配（params["use_regex"]=True）
  - 惰性加载：首次 search() 时触发，connect() 可显式触发

config 结构示例：
  type: excel_file
  enabled: true
  has_header: true
  sheet: 0                  # 工作表索引（0起）或工作表名称
  max_rows: 100000
  sources:
    - type: file
      path: C:/data/materials.xlsx
    - type: directory
      path: C:/data/reports/
      recursive: false
  columns:
    - col: A
      field: 物料代码
    - col: B
      field: 物资名称

search() 的 fields：{"物料代码": "C123", "物资名称": "螺栓"}
search() 的 params：limit(int)、use_regex(bool)
"""

from __future__ import annotations

import logging
import re as _re
import time
from pathlib import Path
from typing import Any

from core.query_base import QueryBase, register_backend
from core.i18n import _t

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════
# 读取引擎检测（模块级，只运行一次）
# ══════════════════════════════════════════════════════════════════

_HAS_CALAMINE = False
_HAS_OPENPYXL = False

try:
    from python_calamine import CalamineWorkbook as _CalamineWorkbook
    _HAS_CALAMINE = True
    logger.debug("excel_file 后端：使用 python-calamine（Rust 引擎）")
except ImportError:
    pass

try:
    import openpyxl as _openpyxl
    _HAS_OPENPYXL = True
    if not _HAS_CALAMINE:
        logger.warning(
            "excel_file 后端：python-calamine 未安装，已回退到 openpyxl（较慢）。"
            " 建议运行: pip install python-calamine"
        )
except ImportError:
    pass

if not _HAS_CALAMINE and not _HAS_OPENPYXL:
    logger.error(
        "excel_file 后端：未找到任何可用的 Excel 读取库。"
        " 请运行: pip install python-calamine"
    )

# calamine 额外支持 .xls 和 .ods
_EXCEL_SUFFIXES: frozenset[str] = (
    frozenset({".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".ods"})
    if _HAS_CALAMINE
    else frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})
)


# ══════════════════════════════════════════════════════════════════
# 工具函数（无外部依赖）
# ══════════════════════════════════════════════════════════════════

def _col_id_to_index(col_str: str) -> int:
    """
    列标识符 → 0-based 整数列索引（纯 Python，无需 openpyxl）。

    支持：
      - 列字母  "A" → 0,  "B" → 1,  "AA" → 26
      - 1-based 数字字符串  "1" → 0,  "2" → 1

    返回 -1 表示无法解析。
    """
    col_str = col_str.strip()
    if not col_str:
        return -1
    if col_str.isdigit():
        n = int(col_str)
        return n - 1 if n >= 1 else -1
    if col_str.isalpha():
        result = 0
        for ch in col_str.upper():
            result = result * 26 + (ord(ch) - 64)   # A=1
        return result - 1
    return -1


def _index_to_col_letter(idx: int) -> str:
    """
    0-based 整数 → Excel 列字母。
    0 → "A",  25 → "Z",  26 → "AA"
    """
    result = ""
    n = idx + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _cell_to_str(v: Any) -> str:
    """
    将 calamine / openpyxl 读取到的单元格值统一转换为字符串。
    calamine 可能返回 int, float, bool, datetime, date, time, str, None。
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "True" if v else "False"
    # datetime / date / time → ISO 字符串
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    # float：去掉多余的小数零（如 1.0 → "1"）
    if isinstance(v, float):
        return f"{v:g}"
    return str(v).strip()


def _match_filter_val(val_str: str, filter_val: Any, suffix: str) -> bool:
    """对 Excel 单元格字符串值进行过滤条件匹配"""
    if not val_str:
        return False
    
    # 1. 最小值 / 起始条件
    if suffix in ("_min", "_from"):
        try:
            return float(val_str) >= float(filter_val)
        except ValueError:
            pass
        return val_str >= str(filter_val)
        
    # 2. 最大值 / 截止条件
    elif suffix in ("_max", "_to"):
        try:
            return float(val_str) <= float(filter_val)
        except ValueError:
            pass
        return val_str <= str(filter_val)
        
    # 3. 精确匹配 (如 checkbox, dropdown 等)
    else:
        if isinstance(filter_val, bool):
            val_bool = val_str.lower() in ("true", "1", "yes", "是")
            return val_bool == filter_val
        return val_str.lower() == str(filter_val).lower()


# ══════════════════════════════════════════════════════════════════

@register_backend("excel_file")
class ExcelFileQuery(QueryBase):
    """
    从本地 Excel 文件读取数据并在内存中执行检索。
    优先使用 python-calamine（Rust，高性能），降级到 openpyxl（纯 Python）。
    """

    def __init__(
        self,
        sources:    list[dict],
        has_header: bool,
        sheet:      int | str,
        columns:    list[dict],
        max_rows:   int = 100_000,
        timeout:    int = 30,         # 保留字段，暂不使用
    ):
        self._sources    = sources    # [{"type": "file"|"directory", "path": "...", "recursive": bool}]
        self._has_header = has_header
        self._sheet      = sheet     # int（0-based 索引）或 str（工作表名）
        self._columns    = columns   # [{"col": "A", "field": "物料代码"}, ...]
        self._max_rows   = max_rows
        self._timeout    = timeout

        self._data: list[dict]              = []
        self._discovered_fields: list[tuple] = []   # [(field_name, display_name)]
        self._status     = "not_loaded"
        self._load_error = ""

    # ── 工厂方法 ────────────────────────────────────────────────

    @classmethod
    def from_config(cls, config: dict[str, Any]) -> "ExcelFileQuery":
        return cls(
            sources    = config.get("sources", []),
            has_header = bool(config.get("has_header", True)),
            sheet      = config.get("sheet", 0),
            columns    = config.get("columns", []),
            max_rows   = int(config.get("max_rows", 100_000)),
            timeout    = int(config.get("timeout", 30)),
        )

    # ── 连接 / 加载 ─────────────────────────────────────────────

    def connect(self) -> bool:
        """显式加载所有数据源到内存，成功返回 True。"""
        if not _HAS_CALAMINE and not _HAS_OPENPYXL:
            self._status = "error"
            self._load_error = (
                "缺少 Excel 读取库，请运行: pip install python-calamine"
            )
            logger.error(self._load_error)
            return False

        try:
            self._data = []
            self._discovered_fields = []
            self._load_error = ""

            files = self._collect_files()
            if not files:
                logger.warning("Excel 文件查询：未发现任何可读取的 Excel 文件")

            engine = "calamine" if _HAS_CALAMINE else "openpyxl"
            for f in files:
                try:
                    t0   = time.perf_counter()
                    rows = self._load_file(f)
                    cost = round(time.perf_counter() - t0, 2)
                    self._data.extend(rows)
                    logger.info(
                        "Excel[%s] 已加载: %s  %d 行  %.2fs",
                        engine, f.name, len(rows), cost,
                    )
                except Exception as exc:
                    logger.error("加载 Excel 文件失败 [%s]: %s", f, exc)

            self._status = "ready"
            logger.info(
                "Excel 后端就绪（引擎=%s）  总行=%d  字段=%s",
                engine, len(self._data),
                [f for f, _ in self._discovered_fields],
            )
            return True

        except Exception as exc:
            self._status = "error"
            self._load_error = str(exc)
            logger.error("Excel 后端加载失败: %s", exc)
            return False

    # ── 文件收集 ────────────────────────────────────────────────

    def _collect_files(self) -> list[Path]:
        """从 sources 配置中收集所有有效的 Excel 文件路径（去重、保序）。"""
        seen:   set[Path]  = set()
        result: list[Path] = []

        for src in self._sources:
            src_type = src.get("type", "file")
            raw_path = src.get("path", "")
            if not raw_path:
                continue
            p = Path(raw_path)

            if src_type == "file":
                if p.is_file() and p.suffix.lower() in _EXCEL_SUFFIXES:
                    if p not in seen:
                        seen.add(p)
                        result.append(p)
                else:
                    logger.warning("Excel 源文件不存在或格式不支持: %s", p)

            elif src_type == "directory":
                if not p.is_dir():
                    logger.warning("Excel 源目录不存在: %s", p)
                    continue
                pattern = "**/*" if src.get("recursive", False) else "*"
                for fp in sorted(p.glob(pattern)):
                    if fp.is_file() and fp.suffix.lower() in _EXCEL_SUFFIXES:
                        if fp not in seen:
                            seen.add(fp)
                            result.append(fp)

        return result

    def _get_target_sheets(self, sheet_names: list[str]) -> list[str | int]:
        """根据 self._sheet 配置解析出需要检索的工作表标识列表（工作表名或 0-based 索引）"""
        raw_sheet = self._sheet
        
        # 1. 如果是列表或元组，直接返回
        if isinstance(raw_sheet, (list, tuple)):
            return list(raw_sheet)
            
        # 2. 如果是字符串，支持逗号分隔或 "all" 特殊值
        if isinstance(raw_sheet, str):
            s_str = raw_sheet.strip()
            if s_str.lower() in ("all", "全部", "*"):
                return list(sheet_names)
            if "," in s_str or "，" in s_str:
                # 兼容中英文逗号
                parts = s_str.replace("，", ",").split(",")
                res: list[str | int] = []
                for p in parts:
                    p_strip = p.strip()
                    if p_strip.isdigit():
                        res.append(int(p_strip))
                    elif p_strip:
                        res.append(p_strip)
                return res
            # 单个字符串
            if s_str.isdigit():
                return [int(s_str)]
            return [s_str] if s_str else [0]
            
        # 3. 单个数字
        if isinstance(raw_sheet, int):
            return [raw_sheet]
            
        return [0]

    # ── 单文件加载（引擎分派） ────────────────────────────────────

    def _load_file(self, path: Path) -> list[dict]:
        """读取单个 Excel 文件（支持多工作表），返回行字典列表。"""
        results: list[dict] = []
        
        # 1. 获取所有的工作表名称（用于匹配和 Wildcards）
        if _HAS_CALAMINE:
            wb = _CalamineWorkbook.from_path(str(path))
            sheet_names = wb.sheet_names
        elif _HAS_OPENPYXL:
            wb = _openpyxl.load_workbook(str(path), read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        else:
            return []
            
        # 2. 解析出需要加载的 sheets
        target_sheets = self._get_target_sheets(sheet_names)
        
        # 3. 循环加载每个工作表
        for ts in target_sheets:
            try:
                if _HAS_CALAMINE:
                    raw_rows = self._read_calamine_sheet(path, ts, sheet_names)
                else:
                    raw_rows = self._read_openpyxl_sheet(path, ts, sheet_names)
                
                # 对每个 sheet 独立解析列映射并记录
                sheet_records = self._build_records(raw_rows)
                results.extend(sheet_records)
            except Exception as e:
                logger.error("加载工作表「%s」失败: %s", ts, e)
                
        return results

    # ── calamine 读取器（Rust，推荐） ─────────────────────────────

    def _read_calamine_sheet(self, path: Path, sheet_ident: str | int, sheet_names: list[str]) -> list[tuple[str, ...]]:
        """
        使用 python-calamine 读取特定工作表，返回字符串 tuple 列表。
        """
        wb = _CalamineWorkbook.from_path(str(path))
        # 选择工作表
        if isinstance(sheet_ident, int):
            idx = max(0, min(sheet_ident, len(sheet_names) - 1))
            ws = wb.get_sheet_by_index(idx)
        elif isinstance(sheet_ident, str) and sheet_ident in sheet_names:
            ws = wb.get_sheet_by_name(sheet_ident)
        else:
            # 尝试通过字符串表示的数字匹配
            try:
                idx = int(sheet_ident)
                idx = max(0, min(idx, len(sheet_names) - 1))
                ws = wb.get_sheet_by_index(idx)
            except ValueError:
                ws = wb.get_sheet_by_index(0)

        max_fetch = self._max_rows + (1 if self._has_header else 0)
        raw = ws.to_python(skip_empty_area=False)

        return [
            tuple(_cell_to_str(v) for v in row)
            for row in raw[:max_fetch]
        ]

    # ── openpyxl 读取器（纯 Python，兜底） ───────────────────────

    def _read_openpyxl_sheet(self, path: Path, sheet_ident: str | int, sheet_names: list[str]) -> list[tuple[str, ...]]:
        """
        使用 openpyxl（read_only 模式）读取特定工作表。
        """
        wb = _openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        # 选择工作表
        if isinstance(sheet_ident, int):
            idx = max(0, min(sheet_ident, len(wb.worksheets) - 1))
            ws = wb.worksheets[idx]
        elif isinstance(sheet_ident, str) and sheet_ident in wb.sheetnames:
            ws = wb[sheet_ident]
        else:
            try:
                idx = int(sheet_ident)
                idx = max(0, min(idx, len(wb.worksheets) - 1))
                ws = wb.worksheets[idx]
            except ValueError:
                ws = wb.active

        max_fetch = self._max_rows + (1 if self._has_header else 0)
        result: list[tuple[str, ...]] = []
        for row in ws.iter_rows(values_only=True):
            result.append(tuple(_cell_to_str(v) for v in row))
            if len(result) >= max_fetch:
                break
        wb.close()
        return result

    # ── 公共：列映射 + 记录构建 ──────────────────────────────────

    def _build_records(self, all_rows: list[tuple[str, ...]]) -> list[dict]:
        """
        根据 self._columns 配置，将原始行列表转换为字段字典列表。
        不依赖具体读取引擎，calamine / openpyxl 共用此逻辑。
        """
        if not all_rows:
            return []

        n_cols = max(len(r) for r in all_rows)

        # ── 建立列映射 [(col_0based_index, field_name)] ──────────
        col_map: list[tuple[int, str]] = []

        if self._columns:
            header_row   = all_rows[0] if self._has_header else None
            header_lower = [v.lower() for v in header_row] if header_row else []

            for col_cfg in self._columns:
                col_str = str(col_cfg.get("col", "")).strip()
                field   = str(col_cfg.get("field", col_str)).strip() or col_str

                idx = _col_id_to_index(col_str)

                if idx == -1:
                    # 按表头名查找
                    if header_lower:
                        try:
                            idx = header_lower.index(col_str.lower())
                        except ValueError:
                            logger.warning("Excel 列「%s」在表头中未找到，跳过", col_str)
                            continue
                    else:
                        logger.warning("无效列标识「%s」，跳过", col_str)
                        continue

                if 0 <= idx < n_cols:
                    col_map.append((idx, field))
                else:
                    logger.warning(
                        "列索引 %d（%s）超出范围（共 %d 列），跳过",
                        idx, col_str, n_cols,
                    )
        else:
            # 无配置：自动生成
            if self._has_header and all_rows:
                col_map = [
                    (i, v if v else _index_to_col_letter(i))
                    for i, v in enumerate(all_rows[0])
                ]
            else:
                col_map = [(i, _index_to_col_letter(i)) for i in range(n_cols)]

        # 更新已发现字段（仅首次赋值）
        if not self._discovered_fields:
            self._discovered_fields = [(field, field) for _, field in col_map]

        # 跳过表头行，构建记录
        data_rows = all_rows[1:] if self._has_header else all_rows
        results: list[dict] = []
        for row in data_rows:
            record: dict = {}
            for idx, field in col_map:
                record[field] = row[idx] if idx < len(row) else ""
            results.append(record)

        return results

    # ── 查询接口 ────────────────────────────────────────────────

    def search(
        self,
        fields: dict[str, str],
        params: dict[str, Any],
    ) -> tuple[list[dict], float]:
        """
        在内存数据中检索匹配行。

        fields : 查询条件 {field_name: value}，空值字段跳过
        params : limit(int, 默认 500)、use_regex(bool, 默认 False)
        返回   : (结果列表, 查询耗时秒数)
        """
        # 惰性加载
        if self._status == "not_loaded":
            self.connect()
        if self._status == "error":
            raise RuntimeError(_t("Excel 数据加载失败: {err}").format(err=self._load_error or _t("未知错误")))

        use_regex = bool(params.get("use_regex", False))
        limit     = min(int(params.get("limit", 500)), 50_000)

        active = {k: str(v).strip() for k, v in fields.items() if v and str(v).strip()}
        filters = params.get("filters") or {}
        if not active and not filters:
            return self._data[:limit], 0.0

        t0 = time.perf_counter()
        results: list[dict] = []

        if use_regex:
            patterns: dict = {}
            for k, v in active.items():
                try:
                    patterns[k] = _re.compile(v, _re.IGNORECASE)
                except _re.error as e:
                    raise RuntimeError(_t("正则表达式错误 [{val}]: {err}").format(val=v, err=e)) from e

            for record in self._data:
                if all(
                    pat.search(str(record.get(k, "")))
                    for k, pat in patterns.items()
                ):
                    # 检查 filters 匹配
                    if filters:
                        match_f = True
                        for k, v in filters.items():
                            base_field = k
                            suffix = ""
                            for sfx in ("_min", "_max", "_from", "_to"):
                                if k.endswith(sfx):
                                    base_field = k[:-len(sfx)]
                                    suffix = sfx
                                    break
                            val_str = str(record.get(base_field, "")).strip()
                            if not _match_filter_val(val_str, v, suffix):
                                match_f = False
                                break
                        if not match_f:
                            continue

                    results.append(record)
                    if len(results) >= limit:
                        break
        else:
            # 大小写不敏感匹配（支持通配符）
            patterns: dict = {}
            literal_matches: dict = {}
            for k, v in active.items():
                has_wildcards = any(c in v for c in ('*', '?', '%', '_'))
                if has_wildcards:
                    escaped = _re.escape(v)
                    pat_str = escaped.replace(r'\*', '.*').replace(r'\?', '.').replace('%', '.*').replace('_', '.')
                    try:
                        patterns[k] = _re.compile(pat_str, _re.IGNORECASE)
                    except _re.error as e:
                        raise RuntimeError(_t("通配符解析错误 [{val}]: {err}").format(val=v, err=e)) from e
                else:
                    literal_matches[k] = v.lower()

            for record in self._data:
                match = True
                # 检查字面值匹配
                for k, v_lower in literal_matches.items():
                    if v_lower not in str(record.get(k, "")).lower():
                        match = False
                        break
                if not match:
                    continue
                # 检查通配符匹配
                if patterns:
                    for k, pat in patterns.items():
                        if not pat.search(str(record.get(k, ""))):
                            match = False
                            break
                if not match:
                    continue
                # 检查 filters 匹配
                if filters:
                    for k, v in filters.items():
                        base_field = k
                        suffix = ""
                        for sfx in ("_min", "_max", "_from", "_to"):
                            if k.endswith(sfx):
                                base_field = k[:-len(sfx)]
                                suffix = sfx
                                break
                        val_str = str(record.get(base_field, "")).strip()
                        if not _match_filter_val(val_str, v, suffix):
                            match = False
                            break
                if match:
                    results.append(record)
                    if len(results) >= limit:
                        break

        elapsed = round(time.perf_counter() - t0, 3)
        logger.info(
            "Excel 查询: 条件=%s  匹配=%d / 总行=%d  耗时=%.3fs",
            list(active.keys()), len(results), len(self._data), elapsed,
        )
        return results, elapsed

    # ── 状态与字段 ──────────────────────────────────────────────

    def get_status(self) -> str:
        return self._status

    def get_available_fields(self) -> list[tuple[str, str]]:
        """
        返回字段列表 [(field_name, display_name)]。
        未加载时从列配置推断；加载后返回实际发现的字段。
        """
        if self._discovered_fields:
            return self._discovered_fields.copy()
        if self._columns:
            return [
                (
                    str(c.get("field", c.get("col", ""))),
                    str(c.get("field", c.get("col", ""))),
                )
                for c in self._columns
                if c.get("field") or c.get("col")
            ]
        return []

    def reload(self) -> bool:
        """强制重新加载所有数据（数据源文件更新后调用）。"""
        self._status = "not_loaded"
        self._data = []
        self._discovered_fields = []
        return self.connect()
